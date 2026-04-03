"""Generate full-text screening workbook from included_for_coding.csv.

Creates an Excel workbook with paper metadata pre-filled for reviewers
to record full-text screening decisions.
"""
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

INCLUDED_CSV = Path("05_screening/included_for_coding.csv")
MASTER_CSV = Path("04_deduped_library/master_records.csv")
OUTPUT_XLSX = Path("05_screening/full_text_screening.xlsx")
OUTPUT_CSV = Path("05_screening/full_text_decisions.csv")

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)


def main():
    # Load master records for metadata
    master = {}
    with open(MASTER_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            pid = row.get("paper_id", "").strip()
            master[pid] = {
                "title": row.get("title", ""),
                "authors": row.get("authors", ""),
                "year": row.get("year", ""),
                "venue": row.get("venue", ""),
                "doi": row.get("doi", ""),
            }

    # Load included paper IDs
    included = []
    with open(INCLUDED_CSV, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            included.append(row["paper_id"].strip())

    print("Building full-text screening workbook for %d papers..." % len(included))

    # ── Create Excel workbook ──
    wb = Workbook()

    # -- Screening sheet --
    ws = wb.active
    ws.title = "Screening"
    headers = [
        "paper_id", "title", "authors", "year", "venue", "doi",
        "pdf_available", "decision_reviewer_A", "decision_reviewer_B",
        "conflict", "final_decision", "exclusion_reason", "tier2_applicable", "notes",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, pid in enumerate(included, 2):
        meta = master.get(pid, {})
        ws.cell(row=row_idx, column=1, value=pid)
        ws.cell(row=row_idx, column=2, value=meta.get("title", ""))
        ws.cell(row=row_idx, column=3, value=meta.get("authors", ""))
        ws.cell(row=row_idx, column=4, value=meta.get("year", ""))
        ws.cell(row=row_idx, column=5, value=meta.get("venue", ""))
        ws.cell(row=row_idx, column=6, value=meta.get("doi", ""))
        # pdf_available, decisions, etc. left blank for reviewers

    # Set column widths
    widths = [14, 60, 40, 6, 30, 30, 14, 20, 20, 10, 15, 16, 16, 40]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w

    # -- Instructions sheet --
    ws_inst = wb.create_sheet("Instructions")
    instructions = [
        ["Full-Text Screening Instructions"],
        [""],
        ["1. Check PDF availability — mark 'yes' or 'no' in pdf_available"],
        ["2. Read each paper's full text"],
        ["3. Record your decision: 'include' or 'exclude'"],
        ["4. If excluding, use a reason code from exclusion_reason_codes.md:"],
        ["   EX-PARADIGM, EX-NONFIN, EX-NOMETHOD, EX-NOEVAL, EX-NOWORKLOAD,"],
        ["   EX-TOOSHORT, EX-DUP, EX-NOACCESS, EX-NOTEN, EX-OTHER"],
        ["5. tier2_applicable: 'yes' if paper has quantitative evaluation relevant to advantage"],
        ["6. Conflicts are resolved by discussion"],
        [""],
        ["Total papers to screen: %d" % len(included)],
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, val in enumerate(row_data, 1):
            ws_inst.cell(row=row_idx, column=col_idx, value=val)
    ws_inst.column_dimensions["A"].width = 80

    wb.save(OUTPUT_XLSX)
    print("Wrote: %s" % OUTPUT_XLSX)

    # ── Also create the CSV template ──
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        csv_headers = [
            "paper_id", "decision_reviewer_A", "decision_reviewer_B",
            "conflict", "final_decision", "exclusion_reason",
            "tier2_applicable", "notes",
        ]
        writer.writerow(csv_headers)
        for pid in included:
            writer.writerow([pid] + [""] * 7)

    print("Wrote: %s" % OUTPUT_CSV)
    print("Done — %d papers ready for full-text screening." % len(included))


if __name__ == "__main__":
    main()
