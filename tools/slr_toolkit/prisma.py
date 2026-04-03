"""PRISMA flow-diagram counts from screening decision files."""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook

from . import config
from .utils import (
    XLSX_HEADER_FILL,
    XLSX_HEADER_FONT,
    cohens_kappa,
    ensure_dir,
    percent_agreement,
    style_xlsx_header,
)

log = logging.getLogger("slr_toolkit.prisma")

_MISSING = "MISSING_INPUT"


def _read_csv_safe(path: Path) -> pd.DataFrame | None:
    """Return DataFrame or None if file missing."""
    if not path.exists():
        log.warning("File not found: %s", path)
        return None
    try:
        return pd.read_csv(path, dtype=str).fillna("")
    except Exception as exc:
        log.error("Error reading %s: %s", path, exc)
        return None


def generate_prisma_counts() -> dict[str, str | int]:
    """Compute PRISMA counts and write to prisma_counts.xlsx.

    Returns the counts dict.
    """
    counts: dict[str, str | int] = {}

    # --- Identification / deduplication -------------------------------------
    master_path = config.MASTER_RECORDS_CSV
    if master_path.exists():
        master = pd.read_csv(master_path, dtype=str).fillna("")
        total_identified = len(master)
        duplicates_removed = int((master["duplicate_of"] != "").sum())
        after_dedup = total_identified - duplicates_removed
    else:
        log.warning("master_records.csv not found — run build-master first.")
        total_identified = _MISSING
        duplicates_removed = _MISSING
        after_dedup = _MISSING

    counts["Identified"] = total_identified
    counts["DuplicatesRemoved"] = duplicates_removed

    # --- Title / abstract screening -----------------------------------------
    # All unique records were screened (AI + human pipeline).  Derive the
    # excluded count from unique − included so the PRISMA flow is consistent.
    screened_ta = after_dedup if isinstance(after_dedup, int) else _MISSING

    # Read the full-text decisions early so we can derive excluded_ta.
    ft_df = _read_csv_safe(config.FT_DECISIONS_FILE)
    if ft_df is not None and "final_decision" in ft_df.columns:
        _included = int(
            (ft_df["final_decision"].str.strip().str.lower() == "include").sum()
        )
        if isinstance(screened_ta, int):
            excluded_ta: int | str = screened_ta - _included
        else:
            excluded_ta = _MISSING
    else:
        excluded_ta = _MISSING

    counts["ScreenedTitleAbstract"] = screened_ta
    counts["ExcludedTitleAbstract"] = excluded_ta

    # --- Full-text screening ------------------------------------------------
    # ft_df already loaded above for the excluded_ta derivation.
    if ft_df is not None and "final_decision" in ft_df.columns:
        assessed_ft = len(ft_df)
        excluded_ft = (ft_df["final_decision"].str.strip().str.lower() == "exclude").sum()
        included_total = (ft_df["final_decision"].str.strip().str.lower() == "include").sum()

        # Tier 2 applicability — look for a 'tier2_applicable' column
        if "tier2_applicable" in ft_df.columns:
            tier2_applicable = (
                (ft_df["final_decision"].str.lower() == "include")
                & (ft_df["tier2_applicable"].str.strip().str.lower() == "yes")
            ).sum()
        else:
            tier2_applicable = "N/A (add 'tier2_applicable' column to full_text_decisions.csv)"
    else:
        assessed_ft = _MISSING
        excluded_ft = _MISSING
        included_total = _MISSING
        tier2_applicable = _MISSING

    counts["FullTextAssessed"] = assessed_ft
    counts["ExcludedFullText"] = excluded_ft
    counts["IncludedTotal"] = included_total
    counts["Tier2Applicable"] = tier2_applicable

    # --- Snowballing --------------------------------------------------------
    snowball_path = config.SEARCH_LOGS_DIR / "snowball_log.csv"
    snowball_identified = 0
    snowball_included = 0
    if snowball_path.exists():
        sb_df = _read_csv_safe(snowball_path)
        if sb_df is not None and len(sb_df) > 0:
            snowball_identified = len(sb_df)
            if "screened_decision" in sb_df.columns:
                snowball_included = (
                    sb_df["screened_decision"].str.strip().str.lower() == "include"
                ).sum()
    counts["SnowballIdentified"] = snowball_identified
    counts["SnowballIncluded"] = snowball_included

    # --- AI-assisted screening (PRISMA-trAIce) ------------------------------
    ai_disc_path = config.AI_DISCREPANCY_REVIEW
    ai_flagged = 0
    ai_rescued = 0
    if ai_disc_path.exists():
        ai_df = _read_csv_safe(ai_disc_path)
        if ai_df is not None and "discrepancy_type" in ai_df.columns:
            ai_flagged = (ai_df["discrepancy_type"] == "ai_rescue").sum()
            # Records where re-review resulted in inclusion
            if "re_review_decision" in ai_df.columns:
                rescued_mask = (
                    (ai_df["discrepancy_type"] == "ai_rescue")
                    & (ai_df["re_review_decision"].str.strip().str.lower() == "include")
                )
                ai_rescued = rescued_mask.sum()
    counts["AIFlagged"] = ai_flagged
    counts["AIRescued"] = ai_rescued

    # --- Full-text exclusion reason breakdown (PRISMA 2020 §13b) ------------
    reason_counts: dict[str, int] = {}
    if ft_df is not None and "exclusion_reason" in ft_df.columns:
        excluded_rows = ft_df[ft_df["final_decision"].str.strip().str.lower() == "exclude"]
        missing_reason = (excluded_rows["exclusion_reason"].str.strip() == "").sum()
        if missing_reason > 0:
            log.warning(
                "%d excluded full-text records have NO exclusion_reason — "
                "please fill in using codes from 05_screening/exclusion_reason_codes.md",
                missing_reason,
            )
        for reason, grp in excluded_rows.groupby(
            excluded_rows["exclusion_reason"].str.strip().str.upper()
        ):
            if reason:
                reason_counts[reason] = len(grp)
        if missing_reason > 0:
            reason_counts["(MISSING REASON)"] = int(missing_reason)

    # --- Calibration metrics ------------------------------------------------
    calibration_metrics: dict[str, str | float | int] | None = None
    cal_df = _read_csv_safe(config.CALIBRATION_DECISIONS_CSV)
    if cal_df is not None and len(cal_df) > 0:
        # Support legacy and current calibration CSV headers.
        if "decision_reviewer_A" in cal_df.columns:
            col_a = "decision_reviewer_A"
        elif "reviewer_a_decision" in cal_df.columns:
            col_a = "reviewer_a_decision"
        else:
            col_a = "decision_A"

        if "decision_reviewer_B" in cal_df.columns:
            col_b = "decision_reviewer_B"
        elif "reviewer_b_decision" in cal_df.columns:
            col_b = "reviewer_b_decision"
        else:
            col_b = "decision_B"
        if col_a in cal_df.columns and col_b in cal_df.columns:
            valid = cal_df[
                (cal_df[col_a].str.strip() != "")
                & (cal_df[col_b].str.strip() != "")
            ]
            if len(valid) > 0:
                dec_a = valid[col_a].str.strip().str.lower().tolist()
                dec_b = valid[col_b].str.strip().str.lower().tolist()
                kappa = cohens_kappa(dec_a, dec_b)
                pct = percent_agreement(dec_a, dec_b)
                calibration_metrics = {
                    "Sample size": len(valid),
                    "Percent agreement": round(pct, 1),
                    "Cohen's kappa": round(kappa, 3),
                    "Target met (kappa >= 0.70)": "Yes" if kappa >= 0.70 else "No",
                }
                log.info(
                    "Calibration: n=%d, agreement=%.1f%%, kappa=%.3f",
                    len(valid), pct, kappa,
                )
    elif config.CALIBRATION_SCREENING_XLSX.exists():
        from .screening import _workbook_screening_rows

        cal_rows = _workbook_screening_rows(config.CALIBRATION_SCREENING_XLSX, dual_review=True)
        valid = [
            row for row in cal_rows
            if row["decision_reviewer_A"] and row["decision_reviewer_B"]
        ]
        if valid:
            dec_a = [row["decision_reviewer_A"] for row in valid]
            dec_b = [row["decision_reviewer_B"] for row in valid]
            kappa = cohens_kappa(dec_a, dec_b)
            pct = percent_agreement(dec_a, dec_b)
            calibration_metrics = {
                "Sample size": len(valid),
                "Percent agreement": round(pct, 1),
                "Cohen's kappa": round(kappa, 3),
                "Target met (kappa >= 0.70)": "Yes" if kappa >= 0.70 else "No",
            }

    # --- Write XLSX ---------------------------------------------------------
    _write_prisma_xlsx(counts, reason_counts, calibration_metrics)

    # Print summary
    print("\nPRISMA Counts:")
    for k, v in counts.items():
        print(f"  {k:30s} : {v}")
    if reason_counts:
        print("\nFull-Text Exclusion Reasons:")
        for reason, n in sorted(reason_counts.items()):
            print(f"  {reason:30s} : {n}")
    if calibration_metrics:
        print("\nCalibration Metrics:")
        for k, v in calibration_metrics.items():
            print(f"  {k:30s} : {v}")
    print()

    return counts


def _write_prisma_xlsx(
    counts: dict[str, str | int],
    reason_counts: dict[str, int] | None = None,
    calibration_metrics: dict[str, str | float | int] | None = None,
) -> None:
    """Write counts to a neatly formatted Excel file."""
    path = config.PRISMA_COUNTS_XLSX
    ensure_dir(path.parent)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "PRISMA Counts"

    ws.append(["Metric", "Count"])
    style_xlsx_header(ws, 2)
    ws.freeze_panes = "A2"

    for metric, value in counts.items():
        ws.append([metric, value if not isinstance(value, str) else str(value)])

    # Exclusion reason breakdown sheet
    if reason_counts:
        ws_reasons = wb.create_sheet("Exclusion Reasons")
        ws_reasons.append(["Exclusion Reason Code", "Count"])
        style_xlsx_header(ws_reasons, 2)
        ws_reasons.freeze_panes = "A2"
        for reason, n in sorted(reason_counts.items()):
            ws_reasons.append([reason, n])
        ws_reasons.column_dimensions["A"].width = 35
        ws_reasons.column_dimensions["B"].width = 12

    # Calibration metrics sheet
    if calibration_metrics:
        ws_cal = wb.create_sheet("Calibration")
        ws_cal.append(["Metric", "Value"])
        style_xlsx_header(ws_cal, 2)
        ws_cal.freeze_panes = "A2"
        for metric, value in calibration_metrics.items():
            ws_cal.append([metric, value])
        ws_cal.column_dimensions["A"].width = 35
        ws_cal.column_dimensions["B"].width = 18

    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    wb.save(path)
    log.info("Wrote PRISMA counts → %s", path)
