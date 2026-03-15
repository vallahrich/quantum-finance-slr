"""Generate screening Excel workbooks and compute inter-rater reliability."""

from __future__ import annotations

import csv
import logging
import random
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import config
from .utils import cohens_kappa, load_master_records

log = logging.getLogger(__name__)

_TRUTHY_LABEL_VALUES = {"1", "true", "yes", "include", "included", "relevant"}
_FALSY_LABEL_VALUES = {"0", "false", "no", "exclude", "excluded", "irrelevant"}

# ── Styles ────────────────────────────────────────────────────────────────

_HEADER_FONT = Font(name="Segoe UI", bold=True, size=11, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_FONT = Font(name="Segoe UI", size=10)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
_THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_FILL_LIGHT_BLUE = PatternFill(start_color="DEEAF6", end_color="DEEAF6", fill_type="solid")

_INSTR_FONT = Font(name="Segoe UI", size=11)
_INSTR_BOLD = Font(name="Segoe UI", size=11, bold=True)
_INSTR_TITLE = Font(name="Segoe UI", size=14, bold=True, color="0078D4")
_DUAL_REVIEW_HEADERS = [
    "#", "Paper ID", "Title", "Authors", "Year", "DOI",
    "Abstract", "Source", "Reviewer A Decision",
    "Reviewer B Decision", "Final Decision", "Notes",
]
_SPLIT_REVIEW_HEADERS = [
    "#", "Paper ID", "Title", "Authors", "Year", "DOI",
    "Abstract", "Source", "Decision", "Notes",
]
_DUAL_REVIEW_WIDTHS = {
    "A": 5, "B": 14, "C": 50, "D": 30, "E": 7, "F": 22,
    "G": 70, "H": 12, "I": 18, "J": 18, "K": 16, "L": 30,
}
_SPLIT_REVIEW_WIDTHS = {
    "A": 5, "B": 14, "C": 50, "D": 30, "E": 7, "F": 22,
    "G": 70, "H": 12, "I": 14, "J": 30,
}


# ── Helpers ───────────────────────────────────────────────────────────────

def _load_unique_records() -> list[dict[str, str]]:
    """Load non-duplicate records from master_records.csv."""
    return load_master_records(unique_only=True)


def _make_decision_validation() -> DataValidation:
    dv = DataValidation(
        type="list", formula1='"include,exclude,maybe"', allow_blank=True,
    )
    dv.error = "Please select: include, exclude, or maybe"
    dv.errorTitle = "Invalid decision"
    dv.prompt = "Select your screening decision"
    dv.promptTitle = "Decision"
    return dv


def _populate_record_rows(
    ws,
    records: list[dict[str, str]],
    headers: list[str],
    decision_columns: tuple[int, ...],
    dv: DataValidation,
) -> int:
    for i, header in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=header)
    _style_header(ws, len(headers))
    ws.add_data_validation(dv)

    for row_idx, rec in enumerate(records, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=rec.get("paper_id", ""))
        ws.cell(row=row_idx, column=3, value=rec.get("title", ""))
        ws.cell(row=row_idx, column=4, value=rec.get("authors", ""))
        ws.cell(row=row_idx, column=5, value=rec.get("year", ""))
        ws.cell(row=row_idx, column=6, value=rec.get("doi", ""))
        ws.cell(row=row_idx, column=7, value=rec.get("abstract", ""))
        ws.cell(row=row_idx, column=8, value=rec.get("source_db", ""))
        for col in decision_columns:
            dv.add(ws.cell(row=row_idx, column=col))
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = _CELL_FONT
            cell.alignment = _CELL_ALIGN
            cell.border = _THIN_BORDER

    return len(records) + 1


def _apply_sheet_layout(
    ws,
    *,
    max_row: int,
    widths: dict[str, int],
    filter_range: str,
    decision_columns: tuple[str, ...],
) -> None:
    for col_letter in decision_columns:
        _add_conditional_formatting(ws, col_letter, max_row)

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = filter_range


def _create_screening_workbook(
    records: list[dict[str, str]],
    *,
    output_path: Path,
    instructions_type: str,
    tab_color: str,
    headers: list[str],
    decision_columns: tuple[int, ...],
    decision_letters: tuple[str, ...],
    widths: dict[str, int],
) -> Path:
    wb = Workbook()
    _add_instructions_sheet(wb, instructions_type)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    ws = wb.create_sheet("Screening")
    ws.sheet_properties.tabColor = tab_color
    max_row = _populate_record_rows(
        ws, records, headers, decision_columns, _make_decision_validation(),
    )
    _apply_sheet_layout(
        ws,
        max_row=max_row,
        widths=widths,
        filter_range=f"A1:{get_column_letter(len(headers))}{max_row}",
        decision_columns=decision_letters,
    )
    _add_progress_sheet(wb, "Screening", decision_letters[0], len(records))
    wb.save(output_path)
    return output_path


def _add_instructions_sheet(wb: Workbook, sheet_type: str) -> None:
    """Add an Instructions sheet to the workbook."""
    ws = wb.create_sheet("Instructions", 0)
    ws.sheet_properties.tabColor = "0078D4"

    lines = [
        ("Title/Abstract Screening — Instructions", _INSTR_TITLE),
        ("", _INSTR_FONT),
        ("GOAL", _INSTR_BOLD),
        (
            "Decide whether each paper should be INCLUDED or EXCLUDED based on "
            "its title and abstract. When in doubt, INCLUDE.",
            _INSTR_FONT,
        ),
        ("", _INSTR_FONT),
        ("INCLUSION CRITERIA (all must hold):", _INSTR_BOLD),
        ("  1. Uses or proposes a gate-based quantum computing approach", _INSTR_FONT),
        ("  2. Addresses a financial application or use case", _INSTR_FONT),
        (
            "  3. Contains enough detail to extract: problem family, quantum method, "
            "evaluation type",
            _INSTR_FONT,
        ),
        ("", _INSTR_FONT),
        ("QUICK EXCLUSION RULES:", _INSTR_BOLD),
        ("  - Quantum annealing ONLY (no gate-based component) -> exclude", _INSTR_FONT),
        ("  - Quantum-inspired classical algorithms -> exclude", _INSTR_FONT),
        ("  - Not finance (chemistry, logistics, etc.) -> exclude", _INSTR_FONT),
        ("  - Pure hardware, no application -> exclude", _INSTR_FONT),
        ("  - Survey/review papers -> exclude (used for snowballing only)", _INSTR_FONT),
        ("  - Non-English -> exclude", _INSTR_FONT),
        ("", _INSTR_FONT),
        ("HOW TO SCREEN:", _INSTR_BOLD),
        ("  1. Go to the 'Screening' sheet", _INSTR_FONT),
        ("  2. Read the Title and Abstract for each row", _INSTR_FONT),
        (
            "  3. Click the Decision column and select: include / exclude / maybe",
            _INSTR_FONT,
        ),
        ("  4. Add notes if helpful (especially for 'maybe' decisions)", _INSTR_FONT),
        ("  5. The row will auto-colour: green=include, red=exclude, yellow=maybe", _INSTR_FONT),
        ("", _INSTR_FONT),
        ("WHEN IN DOUBT -> choose 'maybe' (will be discussed jointly)", _INSTR_BOLD),
        ("", _INSTR_FONT),
        ("PROGRESS:", _INSTR_BOLD),
        ("  Check the 'Progress' sheet to see how many you have left.", _INSTR_FONT),
    ]

    if sheet_type == "calibration":
        lines.insert(2, ("This is the CALIBRATION file — both reviewers screen the same 50 records.", _INSTR_BOLD))
        lines.insert(3, ("Fill YOUR column only (Reviewer A Decision or Reviewer B Decision).", _INSTR_FONT))
        lines.insert(4, ("", _INSTR_FONT))

    for i, (text, font) in enumerate(lines, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = font
    ws.column_dimensions["A"].width = 90


def _style_header(ws, col_count: int) -> None:
    """Apply header styling to row 1."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border = _THIN_BORDER


def _add_conditional_formatting(ws, decision_col_letter: str, max_row: int) -> None:
    """Add green/red/yellow conditional formatting on the decision column."""
    rng = f"{decision_col_letter}2:{decision_col_letter}{max_row}"
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"include"'], fill=_FILL_GREEN)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"exclude"'], fill=_FILL_RED)
    )
    ws.conditional_formatting.add(
        rng, CellIsRule(operator="equal", formula=['"maybe"'], fill=_FILL_YELLOW)
    )


def _add_progress_sheet(wb: Workbook, screening_sheet_name: str,
                        decision_col_letter: str, total: int) -> None:
    """Add a Progress tracking sheet with live formulas."""
    ws = wb.create_sheet("Progress")
    ws.sheet_properties.tabColor = "00B050"

    headers = ["Metric", "Count", "Percentage"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        cell.alignment = _HEADER_ALIGN

    sn = screening_sheet_name
    col = decision_col_letter
    rng = f"'{sn}'!{col}2:{col}{total + 1}"

    metrics = [
        ("Total records", str(total), ""),
        ("Screened", f'=COUNTA({rng})', f'=B3/B2'),
        ("Remaining", f'=B2-B3', f'=B4/B2'),
        ("", "", ""),
        ("Include", f'=COUNTIF({rng},"include")', f'=IF(B3>0,B7/B3,"")'),
        ("Exclude", f'=COUNTIF({rng},"exclude")', f'=IF(B3>0,B8/B3,"")'),
        ("Maybe", f'=COUNTIF({rng},"maybe")', f'=IF(B3>0,B9/B3,"")'),
    ]

    for i, (label, formula, pct) in enumerate(metrics, 2):
        ws.cell(row=i, column=1, value=label).font = Font(name="Segoe UI", size=11, bold=bool(label))
        cell_b = ws.cell(row=i, column=2, value=formula if not formula.startswith("=") else None)
        if formula.startswith("="):
            cell_b.value = formula
        cell_b.font = Font(name="Segoe UI", size=14, bold=True)
        cell_b.alignment = Alignment(horizontal="center")

        cell_c = ws.cell(row=i, column=3, value=pct if not pct.startswith("=") else None)
        if pct.startswith("="):
            cell_c.value = pct
        cell_c.number_format = "0.0%"
        cell_c.font = Font(name="Segoe UI", size=11)
        cell_c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14


# ── Main generators ──────────────────────────────────────────────────────

def generate_calibration_workbook(
    records: list[dict[str, str]],
    sample_size: int = 50,
    seed: int = 42,
    output_path: Path | None = None,
) -> Path:
    """Generate the calibration screening Excel (both reviewers)."""
    if output_path is None:
        output_path = config.CALIBRATION_SCREENING_XLSX

    rng = random.Random(seed)
    sample = rng.sample(records, min(sample_size, len(records)))
    _create_screening_workbook(
        sample,
        output_path=output_path,
        instructions_type="calibration",
        tab_color="FFC000",
        headers=_DUAL_REVIEW_HEADERS,
        decision_columns=(9, 10, 11),
        decision_letters=("I", "J", "K"),
        widths=_DUAL_REVIEW_WIDTHS,
    )
    log.info("Calibration workbook: %s (%d records)", output_path, len(sample))
    return output_path


def generate_split_workbooks(
    records: list[dict[str, str]],
    calibration_ids: set[str],
    seed: int = 42,
    output_dir: Path | None = None,
) -> tuple[Path, Path]:
    """Generate Reviewer A and B screening Excels (split)."""
    if output_dir is None:
        output_dir = config.SCREENING_DIR

    # Exclude calibration records
    remaining = [r for r in records if r.get("paper_id", "") not in calibration_ids]

    # Shuffle deterministically then split
    rng = random.Random(seed)
    rng.shuffle(remaining)
    mid = len(remaining) // 2
    split_a = remaining[:mid]
    split_b = remaining[mid:]

    path_a = output_dir / "screening_reviewer_A.xlsx"
    path_b = output_dir / "screening_reviewer_B.xlsx"

    for label, subset, path in [("A", split_a, path_a), ("B", split_b, path_b)]:
        _create_screening_workbook(
            subset,
            output_path=path,
            instructions_type="split",
            tab_color="FFC000",
            headers=_SPLIT_REVIEW_HEADERS,
            decision_columns=(9,),
            decision_letters=("I",),
            widths=_SPLIT_REVIEW_WIDTHS,
        )
        log.info("Reviewer %s workbook: %s (%d records)", label, path, len(subset))

    return path_a, path_b


def generate_screening_excels(
    seed: int = 42,
    validation_size: int = 100,
) -> dict[str, Path]:
    """Generate all screening workbooks (calibration + validation + split)."""
    records = _load_unique_records()
    log.info("Loaded %d unique records for screening", len(records))

    cal_path = generate_calibration_workbook(records, seed=seed)

    # Get calibration paper IDs
    cal_wb = load_workbook(cal_path, read_only=True)
    cal_ws = cal_wb["Screening"]
    cal_ids: set[str] = set()
    for row in cal_ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if row[1]:
            cal_ids.add(str(row[1]))
    cal_wb.close()

    # Generate held-out validation workbook (Protocol §8, Step 3)
    val_path, val_ids = generate_validation_workbook(
        records, cal_ids, sample_size=validation_size, seed=seed,
    )

    excluded_ids = cal_ids | val_ids
    path_a, path_b = generate_split_workbooks(
        records, excluded_ids, seed=seed, output_dir=config.SCREENING_DIR,
    )

    return {
        "calibration": cal_path,
        "validation": val_path,
        "reviewer_a": path_a,
        "reviewer_b": path_b,
    }


# ── Kappa computation ────────────────────────────────────────────────────

def compute_kappa(calibration_path: Path | None = None) -> dict:
    """Compute Cohen's kappa from the calibration workbook.

    Returns dict with: n, agreement, kappa, details.
    """
    if calibration_path is None:
        calibration_path = config.CALIBRATION_SCREENING_XLSX

    wb = load_workbook(calibration_path, read_only=True)
    ws = wb["Screening"]

    pairs: list[tuple[str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] is None:  # no title = empty row
            continue
        a = str(row[8] or "").strip().lower()
        b = str(row[9] or "").strip().lower()
        if a and b:
            pairs.append((a, b))
    wb.close()

    if not pairs:
        return {"n": 0, "agreement": 0, "kappa": 0, "error": "No paired decisions found"}

    n = len(pairs)
    agree = sum(1 for a, b in pairs if a == b)
    po = agree / n  # observed agreement
    kappa = cohens_kappa([a for a, _ in pairs], [b for _, b in pairs])

    # Confusion details
    confusion: dict[tuple[str, str], int] = {}
    for a, b in pairs:
        key = (a, b)
        confusion[key] = confusion.get(key, 0) + 1

    disagreements = [(a, b) for a, b in pairs if a != b]

    return {
        "n": n,
        "agreement": po,
        "kappa": round(kappa, 3),
        "screened_total": n,
        "agreed": agree,
        "disagreed": len(disagreements),
        "confusion": confusion,
        "pass": kappa >= 0.70,
    }


# ── Merge results ────────────────────────────────────────────────────────

def merge_screening_results(
    calibration_path: Path | None = None,
    reviewer_a_path: Path | None = None,
    reviewer_b_path: Path | None = None,
    output_path: Path | None = None,
) -> Path:
    """Merge calibration + split screening results into one decisions CSV."""
    if calibration_path is None:
        calibration_path = config.CALIBRATION_SCREENING_XLSX
    if reviewer_a_path is None:
        reviewer_a_path = config.REVIEWER_A_SCREENING_XLSX
    if reviewer_b_path is None:
        reviewer_b_path = config.REVIEWER_B_SCREENING_XLSX
    if output_path is None:
        output_path = config.TA_DECISIONS_FILE

    rows: list[dict[str, str]] = []

    # Calibration: use final_decision column
    wb = load_workbook(calibration_path, read_only=True)
    ws = wb["Screening"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] is None:
            continue
        decision = str(row[10] or row[8] or "").strip().lower()  # final > reviewer_a
        rows.append({
            "paper_id": str(row[1] or ""),
            "title": str(row[2] or ""),
            "decision": decision,
            "reviewer": "calibration",
            "notes": str(row[11] or ""),
        })
    wb.close()

    # Reviewer A and B
    for label, path in [("A", reviewer_a_path), ("B", reviewer_b_path)]:
        if not path.exists():
            log.warning("Reviewer %s file not found: %s", label, path)
            continue
        wb = load_workbook(path, read_only=True)
        ws = wb["Screening"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is None:
                continue
            decision = str(row[8] or "").strip().lower()
            rows.append({
                "paper_id": str(row[1] or ""),
                "title": str(row[2] or ""),
                "decision": decision,
                "reviewer": f"reviewer_{label}",
                "notes": str(row[9] or ""),
            })
        wb.close()

    fieldnames = ["paper_id", "title", "decision", "reviewer", "notes"]
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    log.info("Merged %d decisions -> %s", len(rows), output_path)
    return output_path


# ── AI-assisted screening (Protocol §8, Amendment A8) ─────────────────────


def export_asreview_dataset(
    output_path: Path | None = None,
    *,
    exclude_ids: set[str] | None = None,
) -> Path:
    """Export master records as a CSV importable by ASReview LAB.

    ASReview expects columns: ``title``, ``abstract``, ``doi``, ``authors``,
    ``year``, and optionally ``label_included`` (1/0) for prior-knowledge
    records.  The ``paper_id`` is preserved as an extra column so results
    can be joined back.

    Parameters
    ----------
    output_path:
        Destination CSV.  Defaults to ``05_screening/asreview_dataset.csv``.
    exclude_ids:
        Paper IDs to omit (e.g. calibration training records that will be
        labelled separately as prior knowledge).
    """
    if output_path is None:
        output_path = config.ASREVIEW_DATASET_CSV
    if exclude_ids is None:
        exclude_ids = set()

    records = _load_unique_records()
    rows: list[dict[str, str]] = []
    for rec in records:
        pid = rec.get("paper_id", "")
        if pid in exclude_ids:
            continue
        rows.append({
            "paper_id": pid,
            "title": rec.get("title", ""),
            "abstract": rec.get("abstract", ""),
            "authors": rec.get("authors", ""),
            "year": rec.get("year", ""),
            "doi": rec.get("doi", ""),
        })

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["paper_id", "title", "abstract", "authors", "year", "doi"],
        )
        writer.writeheader()
        writer.writerows(rows)

    log.info("Exported %d records for ASReview -> %s", len(rows), output_path)
    return output_path


def export_asreview_labels(
    calibration_path: Path | None = None,
    validation_path: Path | None = None,
    output_path: Path | None = None,
) -> Path:
    """Export consensus labels as ASReview prior-knowledge CSV.

    Reads labelled records from the calibration workbook and, if available,
    the validation workbook.  Both are dual-screened subsets whose labels
    can safely be used as prior knowledge for ASReview active learning.

    Writes ``paper_id``, ``title``, ``abstract``, ``doi``, ``label_included``
    where ``label_included = 1`` for include, ``0`` for exclude.  Records
    with ``maybe`` or no final decision are excluded.
    """
    if calibration_path is None:
        calibration_path = config.CALIBRATION_SCREENING_XLSX
    if validation_path is None:
        validation_path = config.VALIDATION_SCREENING_XLSX
    if output_path is None:
        output_path = config.ASREVIEW_PRIOR_LABELS_CSV

    from openpyxl import load_workbook as _load_wb

    def _read_labels_from_workbook(path: Path) -> list[dict[str, str]]:
        """Read labelled rows from an Excel screening workbook."""
        wb = _load_wb(path, read_only=True)
        ws = wb["Screening"]
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[2] is None:
                continue
            # final decision (col K=10) > reviewer A (col I=8) fallback
            decision = str(row[10] or row[8] or "").strip().lower()
            if decision not in ("include", "exclude"):
                continue
            rows.append({
                "paper_id": str(row[1] or ""),
                "title": str(row[2] or ""),
                "abstract": str(row[6] or ""),
                "doi": str(row[5] or ""),
                "label_included": "1" if decision == "include" else "0",
            })
        wb.close()
        return rows

    rows = _read_labels_from_workbook(calibration_path)
    cal_count = len(rows)

    val_count = 0
    if validation_path.exists():
        seen_ids = {r["paper_id"] for r in rows}
        val_rows = _read_labels_from_workbook(validation_path)
        # Avoid duplicates if any paper appears in both workbooks
        for r in val_rows:
            if r["paper_id"] not in seen_ids:
                rows.append(r)
                seen_ids.add(r["paper_id"])
        val_count = len(rows) - cal_count
        log.info("Including %d validation labels as additional priors", val_count)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["paper_id", "title", "abstract", "doi", "label_included"],
        )
        writer.writeheader()
        writer.writerows(rows)

    log.info(
        "Exported %d prior labels (%d calibration + %d validation) -> %s",
        len(rows), cal_count, val_count, output_path,
    )
    return output_path


def generate_validation_workbook(
    records: list[dict[str, str]],
    calibration_ids: set[str],
    sample_size: int = 100,
    seed: int = 42,
    output_path: Path | None = None,
) -> tuple[Path, set[str]]:
    """Generate held-out validation subset workbook (dual-screened).

    Returns (path, set_of_validation_paper_ids).
    """
    if output_path is None:
        output_path = config.VALIDATION_SCREENING_XLSX

    remaining = [r for r in records if r.get("paper_id", "") not in calibration_ids]
    rng = random.Random(seed + 1)  # distinct seed from calibration sample
    sample = rng.sample(remaining, min(sample_size, len(remaining)))
    validation_ids = {r.get("paper_id", "") for r in sample}
    _create_screening_workbook(
        sample,
        output_path=output_path,
        instructions_type="calibration",
        tab_color="7030A0",
        headers=_DUAL_REVIEW_HEADERS,
        decision_columns=(9, 10, 11),
        decision_letters=("I", "J", "K"),
        widths=_DUAL_REVIEW_WIDTHS,
    )
    log.info("Validation workbook: %s (%d records)", output_path, len(sample))
    return output_path, validation_ids


def import_ai_decisions(
    ai_export_path: Path,
    output_path: Path | None = None,
) -> Path:
    """Import AI screening results (e.g. ASReview export) into pipeline CSV.

    Expects input CSV with at least ``paper_id`` (or ``record_id``) and one
    of ``label_included`` (0/1), ``included`` (0/1), or ``label``
    (``relevant``/``irrelevant``).  Optionally ``confidence`` or
    ``proba``.

    Writes standardised ``ai_screening_decisions.csv`` with columns:
    ``paper_id``, ``ai_decision``, ``ai_confidence``.
    """
    import pandas as pd

    if output_path is None:
        output_path = config.AI_SCREENING_DECISIONS

    df = pd.read_csv(ai_export_path, dtype=str).fillna("")

    def _normalize_binary_series(series: pd.Series, column_name: str) -> pd.Series:
        normalized = series.astype(str).str.strip().str.lower()
        invalid_values = sorted({
            value for value in normalized
            if value and value not in _TRUTHY_LABEL_VALUES and value not in _FALSY_LABEL_VALUES
        })
        if invalid_values:
            raise ValueError(
                f"Unsupported values in '{column_name}': {invalid_values}. "
                "Expected truthy/falsy labels such as 1/0, true/false, yes/no, or include/exclude."
            )
        return normalized.map(
            lambda value: "include" if value in _TRUTHY_LABEL_VALUES else "exclude"
        )

    # Resolve paper_id column
    if "paper_id" not in df.columns:
        if "record_id" in df.columns:
            df = df.rename(columns={"record_id": "paper_id"})
        else:
            raise ValueError(
                f"AI export must have 'paper_id' or 'record_id' column. "
                f"Found: {list(df.columns)}"
            )

    # Resolve decision column
    if "label_included" in df.columns:
        df["ai_decision"] = _normalize_binary_series(df["label_included"], "label_included")
    elif "included" in df.columns:
        df["ai_decision"] = _normalize_binary_series(df["included"], "included")
    elif "label" in df.columns:
        df["ai_decision"] = _normalize_binary_series(df["label"], "label")
    else:
        raise ValueError(
            "AI export must have 'label_included', 'included', or 'label' column."
        )

    # Resolve confidence column
    df["ai_confidence"] = ""
    for col in ("confidence", "proba", "probability", "score"):
        if col in df.columns:
            df["ai_confidence"] = df[col].astype(str)
            break

    out = df[["paper_id", "ai_decision", "ai_confidence"]]
    out.to_csv(output_path, index=False)
    log.info("Imported %d AI decisions -> %s", len(out), output_path)
    return output_path


def find_discrepancies(
    human_decisions_path: Path | None = None,
    ai_decisions_path: Path | None = None,
    output_path: Path | None = None,
) -> dict[str, int]:
    """Join human and AI screening decisions to find discrepancies.

    Writes ``ai_discrepancy_review.csv`` with columns:
    ``paper_id``, ``title``, ``human_decision``, ``ai_decision``,
    ``ai_confidence``, ``discrepancy_type``, ``re_review_decision``,
    ``notes``.

    Discrepancy types:
    - ``ai_rescue``:  AI=include, Human=exclude  (primary safety-net catch)
    - ``agree_include``: both include
    - ``agree_exclude``: both exclude
    - ``human_only``: Human=include, AI=exclude  (human stands)

    Returns counts of each discrepancy type.
    """
    import pandas as pd

    if human_decisions_path is None:
        human_decisions_path = config.TA_DECISIONS_FILE
    if ai_decisions_path is None:
        ai_decisions_path = config.AI_SCREENING_DECISIONS
    if output_path is None:
        output_path = config.AI_DISCREPANCY_REVIEW

    human_df = pd.read_csv(human_decisions_path, dtype=str).fillna("")
    ai_df = pd.read_csv(ai_decisions_path, dtype=str).fillna("")

    # Normalise human decision column name
    if "final_decision" in human_df.columns:
        human_df["human_decision"] = human_df["final_decision"].str.strip().str.lower()
    elif "decision" in human_df.columns:
        human_df["human_decision"] = human_df["decision"].str.strip().str.lower()
    else:
        raise ValueError("Human decisions file must have 'final_decision' or 'decision' column.")

    merged = human_df[["paper_id", "human_decision"]].merge(
        ai_df[["paper_id", "ai_decision", "ai_confidence"]],
        on="paper_id",
        how="inner",
    )

    # Add title from master if available
    title_map: dict[str, str] = {}
    if config.MASTER_RECORDS_CSV.exists():
        master = pd.read_csv(config.MASTER_RECORDS_CSV, dtype=str).fillna("")
        title_map = dict(zip(master["paper_id"], master["title"]))
    merged["title"] = merged["paper_id"].map(title_map).fillna("")

    # Classify discrepancy type
    def _classify(row: pd.Series) -> str:
        h = row["human_decision"]
        a = row["ai_decision"]
        if a == "include" and h == "exclude":
            return "ai_rescue"
        if a == "include" and h == "include":
            return "agree_include"
        if a == "exclude" and h == "exclude":
            return "agree_exclude"
        if a == "exclude" and h == "include":
            return "human_only"
        return "other"

    merged["discrepancy_type"] = merged.apply(_classify, axis=1)
    merged["re_review_decision"] = ""
    merged["notes"] = ""

    out_cols = [
        "paper_id", "title", "human_decision", "ai_decision",
        "ai_confidence", "discrepancy_type", "re_review_decision", "notes",
    ]
    merged[out_cols].to_csv(output_path, index=False)

    counts = merged["discrepancy_type"].value_counts().to_dict()
    log.info("Discrepancy analysis -> %s", output_path)
    for dtype, n in sorted(counts.items()):
        log.info("  %s: %d", dtype, n)
    return counts


def generate_fn_audit(
    discrepancy_path: Path | None = None,
    output_path: Path | None = None,
    audit_fraction: float = 0.10,
    seed: int = 42,
) -> Path:
    """Sample 10% of double-excluded records for false-negative audit.

    Reads ``ai_discrepancy_review.csv`` and samples from rows where
    ``discrepancy_type == 'agree_exclude'``.  Writes audit CSV for the
    second reviewer to re-screen.
    """
    import pandas as pd

    if discrepancy_path is None:
        discrepancy_path = config.AI_DISCREPANCY_REVIEW
    if output_path is None:
        output_path = config.FN_AUDIT_SAMPLE

    df = pd.read_csv(discrepancy_path, dtype=str).fillna("")
    double_excluded = df[df["discrepancy_type"] == "agree_exclude"]

    n_sample = max(1, int(len(double_excluded) * audit_fraction))
    sample = double_excluded.sample(n=min(n_sample, len(double_excluded)),
                                    random_state=seed)

    # Add audit columns
    sample = sample.copy()
    sample["audit_decision"] = ""
    sample["audit_notes"] = ""

    out_cols = [
        "paper_id", "title", "human_decision", "ai_decision",
        "ai_confidence", "audit_decision", "audit_notes",
    ]
    sample[out_cols].to_csv(output_path, index=False)
    log.info(
        "FN audit: sampled %d / %d double-excluded records -> %s",
        len(sample), len(double_excluded), output_path,
    )
    return output_path


def compute_ai_validation(
    validation_path: Path | None = None,
    ai_decisions_path: Path | None = None,
    report_path: Path | None = None,
) -> dict[str, float | int | str]:
    """Compute AI performance on the held-out validation subset.

    Reads human consensus decisions from the validation workbook and AI
    decisions from ``ai_screening_decisions.csv``.  Computes recall,
    specificity, precision, F1, Cohen's \u03ba.  Writes a Markdown report.

    Returns dict with metrics and a ``pass`` boolean (recall >= 0.95).
    """
    import pandas as pd
    from openpyxl import load_workbook as _load_wb

    if validation_path is None:
        validation_path = config.VALIDATION_SCREENING_XLSX
    if ai_decisions_path is None:
        ai_decisions_path = config.AI_SCREENING_DECISIONS
    if report_path is None:
        report_path = config.AI_VALIDATION_REPORT

    # Read human consensus from validation workbook
    wb = _load_wb(validation_path, read_only=True)
    ws = wb["Screening"]
    human: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2] is None:
            continue
        pid = str(row[1] or "")
        decision = str(row[10] or row[8] or "").strip().lower()
        if decision in ("include", "exclude"):
            human[pid] = decision
    wb.close()

    # Read AI decisions
    ai_df = pd.read_csv(ai_decisions_path, dtype=str).fillna("")
    ai_map: dict[str, str] = dict(zip(ai_df["paper_id"], ai_df["ai_decision"]))

    # Compute metrics on the intersection
    common_ids = sorted(set(human.keys()) & set(ai_map.keys()))
    if not common_ids:
        return {"error": "No overlapping records between validation and AI decisions"}

    y_true = [1 if human[pid] == "include" else 0 for pid in common_ids]
    y_pred = [1 if ai_map[pid] == "include" else 0 for pid in common_ids]

    tp = sum(t == 1 and p == 1 for t, p in zip(y_true, y_pred))
    fn = sum(t == 1 and p == 0 for t, p in zip(y_true, y_pred))
    fp = sum(t == 0 and p == 1 for t, p in zip(y_true, y_pred))
    tn = sum(t == 0 and p == 0 for t, p in zip(y_true, y_pred))

    recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    specificity = tn / (tn + fp) if (tn + fp) > 0 else 0.0
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    f1 = (2 * precision * recall / (precision + recall)
           if (precision + recall) > 0 else 0.0)

    # Cohen's kappa
    n = len(common_ids)
    po = (tp + tn) / n
    pe = (((tp + fn) * (tp + fp)) + ((tn + fp) * (tn + fn))) / (n * n)
    kappa = (po - pe) / (1 - pe) if pe < 1.0 else 1.0

    metrics = {
        "n": n,
        "tp": tp, "fn": fn, "fp": fp, "tn": tn,
        "recall": round(recall, 4),
        "specificity": round(specificity, 4),
        "precision": round(precision, 4),
        "f1": round(f1, 4),
        "kappa": round(kappa, 4),
        "pass": recall >= 0.95,
    }

    # Write Markdown report
    _pass_fail = "PASS \u2713" if recall >= 0.95 else "FAIL \u2717"
    report = (
        "# AI Validation Report\n\n"
        "## Held-out validation subset performance\n\n"
        f"| Metric | Value |\n"
        f"|--------|-------|\n"
        f"| Records evaluated | {n} |\n"
        f"| True positives (AI=incl, Human=incl) | {tp} |\n"
        f"| False negatives (AI=excl, Human=incl) | {fn} |\n"
        f"| False positives (AI=incl, Human=excl) | {fp} |\n"
        f"| True negatives (AI=excl, Human=excl) | {tn} |\n"
        f"| **Recall (sensitivity)** | **{recall:.4f}** |\n"
        f"| Specificity | {specificity:.4f} |\n"
        f"| Precision | {precision:.4f} |\n"
        f"| F1 score | {f1:.4f} |\n"
        f"| Cohen's \u03ba (human-AI) | {kappa:.4f} |\n\n"
        f"## Threshold check\n\n"
        f"- Required: recall \u2265 0.95\n"
        f"- Achieved: recall = {recall:.4f}\n"
        f"- **Result: {_pass_fail}**\n\n"
    )

    if not metrics["pass"]:
        report += (
            "### Action required\n\n"
            "AI recall is below the 0.95 threshold. Options:\n"
            "1. Refine ASReview model configuration and re-evaluate\n"
            "2. Abandon AI layer and proceed with purely human screening\n\n"
        )

    with open(report_path, "w", encoding="utf-8") as f:
        f.write(report)

    log.info("AI validation report -> %s (recall=%.4f, %s)",
             report_path, recall, "PASS" if metrics["pass"] else "FAIL")
    return metrics


def run_asreview_simulate(
    dataset_path: Path | None = None,
    labels_path: Path | None = None,
    output_path: Path | None = None,
    *,
    model: str = "elas_u4",
    seed: int = 42,
    threshold: float = 0.5,
) -> Path:
    """Run ASReview active-learning ranking and export decisions.

    Uses ASReview's ``ActiveLearningCycle`` to train on prior labels and
    rank all unlabeled records by predicted relevance.

    Parameters
    ----------
    dataset_path:
        CSV with columns ``paper_id``, ``title``, ``abstract`` (from
        ``export_asreview_dataset``).
    labels_path:
        CSV with ``paper_id`` and ``label_included`` columns (from
        ``export_asreview_labels``).
    output_path:
        Destination CSV.  Defaults to ``config.AI_SCREENING_DECISIONS``.
    model:
        ASReview model configuration name (default ``elas_u4``).
    seed:
        Random seed for reproducibility.
    threshold:
        Probability threshold for inclusion decision.  Records with
        P(relevant) >= threshold are marked ``include`` (default 0.5).
    """
    import numpy as np
    import pandas as pd
    from asreview import ActiveLearningCycle
    from asreview.models import get_ai_config

    if dataset_path is None:
        dataset_path = config.ASREVIEW_DATASET_CSV
    if labels_path is None:
        labels_path = config.ASREVIEW_PRIOR_LABELS_CSV
    if output_path is None:
        output_path = config.AI_SCREENING_DECISIONS

    # Load dataset and prior labels
    dataset = pd.read_csv(dataset_path, dtype=str).fillna("")
    priors = pd.read_csv(labels_path, dtype=str).fillna("")

    # Prior records may have been excluded from the dataset during export.
    # Merge them back so the model can train on them.
    prior_ids_in_dataset = set(dataset["paper_id"]) & set(priors["paper_id"])
    missing_priors = priors[~priors["paper_id"].isin(prior_ids_in_dataset)]
    if not missing_priors.empty:
        # Build rows for missing priors using fields available in priors CSV
        extra_rows = []
        for _, row in missing_priors.iterrows():
            extra_rows.append({
                "paper_id": row.get("paper_id", ""),
                "title": row.get("title", ""),
                "abstract": row.get("abstract", ""),
                "authors": row.get("authors", ""),
                "year": row.get("year", ""),
                "doi": row.get("doi", ""),
            })
        dataset = pd.concat(
            [dataset, pd.DataFrame(extra_rows).fillna("")], ignore_index=True,
        )

    # Map prior labels onto the dataset
    prior_map = dict(zip(priors["paper_id"], priors["label_included"].astype(int)))
    dataset["label_included"] = dataset["paper_id"].map(prior_map).fillna(-1).astype(int)

    labels = dataset["label_included"].values.copy()
    prior_mask = labels >= 0
    unlabeled_mask = ~prior_mask

    # Get model config and instantiate model objects
    ai_config = get_ai_config(model)
    cycle_data = ai_config["value"]
    cycle = ActiveLearningCycle.from_meta(cycle_data)

    np.random.seed(seed)

    # ASReview feature extractor expects DataFrame with 'title' and 'abstract'
    X = dataset[["title", "abstract"]].copy()
    X_features = cycle.transform(X)

    # Train on prior-labelled records
    cycle.fit(X_features[prior_mask], labels[prior_mask])

    # Compute relevance scores for unlabeled records
    X_unlabeled = X_features[unlabeled_mask]
    classifier = cycle.classifier

    try:
        proba = classifier.predict_proba(X_unlabeled)
        scores = proba[:, 1]  # P(relevant)
    except AttributeError:
        # SVM / classifiers without predict_proba: use decision_function
        # and normalise to [0, 1] via sigmoid
        raw = classifier.decision_function(X_unlabeled)
        scores = 1.0 / (1.0 + np.exp(-raw))

    # Rank by score descending (rank 1 = most relevant)
    ranked_order = np.argsort(-scores)

    # Map back to dataset positions
    unlabeled_positions = np.where(unlabeled_mask)[0]
    ranked_dataset_indices = unlabeled_positions[ranked_order]

    # Build output rows ranked by AI relevance (rank 1 = most relevant)
    paper_ids = dataset["paper_id"].tolist()
    out_rows = []
    for rank, order_idx in enumerate(ranked_order, 1):
        ds_idx = ranked_dataset_indices[rank - 1]
        pid = paper_ids[ds_idx]
        score = float(scores[order_idx])
        out_rows.append({
            "paper_id": pid,
            "ai_decision": "include" if score >= threshold else "exclude",
            "ai_rank": rank,
            "ai_confidence": f"{score:.4f}",
        })

    out_df = pd.DataFrame(out_rows)
    out_df.to_csv(output_path, index=False)
    n_include = sum(1 for r in out_rows if r["ai_decision"] == "include")
    log.info(
        "ASReview ranking complete: %d records ranked, %d flagged include -> %s",
        len(out_rows), n_include, output_path,
    )
    return output_path
