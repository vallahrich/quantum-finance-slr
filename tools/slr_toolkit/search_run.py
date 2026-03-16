"""Search-run management — create dated run folders, update search log."""

from __future__ import annotations

import logging
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from . import config
from .utils import ensure_dir, style_xlsx_header

log = logging.getLogger("slr_toolkit.search_run")


def _ensure_search_log_xlsx() -> Path:
    """Create search_log.xlsx with headers if it doesn't exist; return path."""
    path = config.SEARCH_LOG_XLSX
    if path.exists():
        return path

    ensure_dir(path.parent)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "SearchLog"
    ws.append(config.SEARCH_LOG_COLUMNS)

    # Style header
    style_xlsx_header(ws, len(config.SEARCH_LOG_COLUMNS))

    # Column widths
    widths = {
        "A": 22, "B": 14, "C": 16, "D": 18,
        "E": 60, "F": 18, "G": 22, "H": 16,
        "I": 18, "J": 12, "K": 14, "L": 30,
        "M": 30,
    }
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w

    ws.freeze_panes = "A2"
    wb.save(path)
    log.info("Created search log: %s", path)
    return path


_RUN_README = """\
Search Run Export Folder
========================
Source database : {source}
Date            : {run_date}

Instructions
------------
1. Run your search query in {source}.
2. Export results in one of these formats:
   - RIS  (.ris)
   - BibTeX (.bib)
   - CSV  (.csv)
3. Place the exported file(s) in THIS folder.
4. Then ingest them:
   python -m tools.slr_toolkit.cli ingest --run-folder "{run_folder}"
5. **PRISMA-S mandatory:** update 02_search_logs/search_log.xlsx with:
   - Interface: e.g. "Scopus web", "WoS Basic Search", "IEEE Xplore"
   - FullSearchString: exact copy-paste of the query as executed
   - Fields: e.g. TITLE-ABS-KEY, ALL
   - DateLimits: e.g. 2016-01-01 to {run_date}
   - LanguageLimits: e.g. English (or blank if none)
   - ResultsN: number of hits returned
"""


def create_search_run(
    source: str,
    run_date: str | None = None,
    *,
    log_search: bool = True,
) -> Path:
    """Create a new search-run folder and append a row to the search log.

    Returns the path to the new run folder.
    """
    if run_date is None:
        run_date = date.today().isoformat()

    run_id = f"{run_date}_{source}"
    run_folder = config.RAW_EXPORTS_DIR / run_id
    ensure_dir(run_folder)

    # Write README.txt inside run folder
    readme_text = _RUN_README.format(
        source=source, run_date=run_date, run_folder=run_folder,
    )
    readme_path = run_folder / "README.txt"
    readme_path.write_text(readme_text, encoding="utf-8")
    log.info("Created run folder: %s", run_folder)

    if not log_search:
        return run_folder

    # Append row to search log
    xlsx_path = _ensure_search_log_xlsx()
    wb = load_workbook(xlsx_path)
    ws = wb.active
    assert ws is not None
    existing_ids = {
        str(value).strip()
        for value, in ws.iter_rows(min_row=2, max_col=1, values_only=True)
        if value
    }
    if run_id in existing_ids:
        log.info("Search run '%s' already present in %s", run_id, xlsx_path)
        wb.close()
        return run_folder

    ws.append([
        run_id,      # SearchRunID
        run_date,    # Date
        source,      # Database
        "",          # Interface  (user fills in, e.g. "Scopus web")
        "",          # FullSearchString  (exact copy-paste)
        "",          # Fields  (e.g. TITLE-ABS-KEY)
        "",          # DateLimits  (e.g. 2016-01-01 to 2026-03-08)
        "",          # LanguageLimits  (e.g. English)
        "",          # OtherLimits
        "",          # ResultsN
        "",          # ExportFormat
        "",          # ExportFiles
        "",          # Notes
    ])
    wb.save(xlsx_path)
    wb.close()
    log.info("Appended search run '%s' to %s", run_id, xlsx_path)

    return run_folder
