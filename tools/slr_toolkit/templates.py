"""Template generators — screening CSVs, extraction XLSX, protocol, codebook."""

from __future__ import annotations

import io
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from . import config
from .utils import ensure_dir, safe_write_bytes, safe_write_text

log = logging.getLogger("slr_toolkit.templates")

# ── Screening templates ────────────────────────────────────────────────────

_TA_HEADER = "paper_id,decision_A,decision_B,conflict,final_decision,reason_code,notes\n"
_FT_HEADER = "paper_id,decision_A,decision_B,conflict,final_decision,exclusion_reason,stage,notes\n"


def create_ta_decisions_template(*, force: bool = False) -> None:
    """Create title/abstract screening template CSV."""
    safe_write_text(config.TA_DECISIONS_TEMPLATE, _TA_HEADER, force=force)


def create_ft_decisions_template(*, force: bool = False) -> None:
    """Create full-text screening template CSV."""
    safe_write_text(config.FT_DECISIONS_TEMPLATE, _FT_HEADER, force=force)


# ── Extraction template ────────────────────────────────────────────────────

_CODEBOOK_ROWS: list[tuple[str, str, str]] = [
    ("paper_id", "Stable hash identifier", "auto-generated"),
    ("title", "Full paper title", "free text"),
    ("authors", "Semicolon-separated author list", "free text"),
    ("year", "Publication year", "integer"),
    ("venue", "Journal / conference / preprint server", "free text"),
    ("doi", "Digital Object Identifier", "DOI string or empty"),
    ("problem_family", "Finance problem addressed",
     "portfolio_optimization | option_pricing | risk_analysis | credit_scoring | "
     "fraud_detection | monte_carlo | time_series | other"),
    ("quantum_method", "Algorithm / approach used",
     "VQE | QAOA | QAE | Grover | HHL | quantum_walk | variational | hybrid_classical | other"),
    ("evaluation_type", "How results were obtained",
     "real_hardware | simulator | analytical | hybrid"),
    ("NISQ_vs_FT", "Target hardware regime",
     "NISQ | fault_tolerant | both | unclear"),
    ("qubit_count", "Number of qubits used/projected", "integer or N/A"),
    ("gate_depth", "Circuit depth (if reported)", "integer or N/A"),
    ("baseline_strength", "Quality of classical baseline",
     "state_of_art | reasonable | weak | none | unclear"),
    ("advantage_claim", "Does the paper claim quantum advantage?",
     "yes | no | projected | unclear"),
    ("advantage_evidence", "Evidence supporting the claim",
     "empirical | analytical | projected | none"),
    ("hardware_or_sim", "Execution environment",
     "ibm | google | ionq | simulator_statevector | simulator_noisy | other | N/A"),
    ("dataset_description", "Data used for evaluation", "free text"),
    # --- Hoefler framework fields (Stage B) ---
    ("input_data_size", "Size/dimensionality of input data to the quantum algorithm",
     "integer or description (e.g., '4 assets', '2^10 grid points') or N/A"),
    ("output_type", "Nature of the quantum algorithm's output",
     "scalar | vector | distribution_sample | expectation_value | other | N/A"),
    ("io_bottleneck_discussed", "Does the paper discuss I/O bandwidth limitations?",
     "yes | no"),
    ("big_compute_small_data", "Does the workload fit the 'big compute on small data' pattern?",
     "yes | no | unclear | N/A"),
    ("speedup_type_detailed", "Speedup characterisation beyond asymptotic class",
     "exponential | quartic | cubic | quadratic | sub-quadratic | none_claimed | unclear"),
    ("speedup_constant_reported", "Are concrete constant factors or prefactors reported (not just big-O)?",
     "yes | no"),
    ("oracle_stateprep_cost_included", "Does the evaluation account for oracle construction and state preparation overhead?",
     "yes | partial | no"),
    ("end_to_end_overhead", "Does the evaluation include full end-to-end overhead (not just query complexity)?",
     "yes | partial | no"),
    ("crossover_time_estimated", "Is a crossover time explicitly estimated?",
     "yes | no"),
    ("crossover_time_value", "Reported crossover time (if estimated)",
     "free text (e.g., '3.2 years', '< 2 weeks', 'not computed') or N/A"),
    ("crossover_size_estimated", "Is a crossover problem size explicitly estimated?",
     "yes | no"),
    ("crossover_size_value", "Reported crossover problem size",
     "free text (e.g., 'N > 10^6 assets', '2^50 grid points') or N/A"),
    ("tier1_achievable", "Could the workload plausibly achieve Tier-1 crossover (≤ 2 weeks)?",
     "yes | no | insufficient_data"),
    ("tier2_finance_sla", "Does the paper assess against a finance-specific operational window?",
     "yes_overnight | yes_intraday | yes_other | no"),
    ("classical_baseline_detail", "Description of classical baseline used for comparison",
     "free text (e.g., 'Monte Carlo on single CPU', 'GPU-accelerated QMC', 'analytical Black-Scholes') or N/A"),
    ("classical_baseline_hardware", "Hardware specification of classical baseline",
     "free text (e.g., 'NVIDIA A100', 'Intel Xeon 48-core', 'unspecified') or N/A"),
    ("qubit_type", "Physical qubit technology assumed or used",
     "superconducting | trapped_ion | photonic | neutral_atom | unspecified | N/A"),
    ("error_correction_model", "Error correction assumptions",
     "surface_code | other_code | error_mitigated_only | noiseless_simulation | unspecified"),
    ("t_count_or_gate_cost", "T-count or dominant gate cost reported",
     "free text with value or N/A"),
    ("shots_or_samples", "Number of measurement shots or samples reported",
     "integer or N/A"),
]

_EXTRACTION_COLUMNS: list[str] = [row[0] for row in _CODEBOOK_ROWS]

_RUBRIC_COLUMNS: list[str] = [
    "paper_id",
    "q_methodology",
    "q_reproducibility",
    "q_baseline",
    "q_scalability",
    "q_justification",
    "q_io_bottleneck",       # I/O limitations acknowledged and addressed
    "q_crossover_framing",   # Tier-1/Tier-2 crossover analysis present
    "q_end_to_end",          # End-to-end overhead included
]

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, ncols: int) -> None:  # type: ignore[no-untyped-def]
    """Apply bold-white-on-blue header styling to the first row."""
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
    ws.freeze_panes = "A2"


def create_extraction_template(*, force: bool = False) -> None:
    """Create the extraction XLSX with Codebook, Extraction, and Rubric sheets."""
    path = config.EXTRACTION_TEMPLATE
    if path.exists() and not force:
        log.info("Skipping (exists): %s", path)
        return

    wb = Workbook()

    # -- Codebook sheet --
    ws_cb = wb.active
    assert ws_cb is not None
    ws_cb.title = "Codebook"
    ws_cb.append(["Column Name", "Definition", "Allowed Values"])
    _style_header(ws_cb, 3)
    for row in _CODEBOOK_ROWS:
        ws_cb.append(list(row))
    ws_cb.column_dimensions["A"].width = 22
    ws_cb.column_dimensions["B"].width = 40
    ws_cb.column_dimensions["C"].width = 60

    # -- Extraction sheet --
    ws_ex = wb.create_sheet("Extraction")
    ws_ex.append(_EXTRACTION_COLUMNS)
    _style_header(ws_ex, len(_EXTRACTION_COLUMNS))
    for i, col_name in enumerate(_EXTRACTION_COLUMNS, start=1):
        ws_ex.column_dimensions[ws_ex.cell(row=1, column=i).column_letter].width = 20

    # -- Rubric sheet --
    ws_rb = wb.create_sheet("Rubric")
    ws_rb.append(_RUBRIC_COLUMNS)
    _style_header(ws_rb, len(_RUBRIC_COLUMNS))
    for i, col_name in enumerate(_RUBRIC_COLUMNS, start=1):
        ws_rb.column_dimensions[ws_rb.cell(row=1, column=i).column_letter].width = 20

    ensure_dir(path.parent)
    buf = io.BytesIO()
    wb.save(buf)
    safe_write_bytes(path, buf.getvalue(), force=True)  # already checked above


# ── Protocol & amendments ──────────────────────────────────────────────────

def create_protocol(*, force: bool = False) -> None:
    """Write protocol skeleton to 01_protocol/protocol_v1.0.md."""
    # Protocol is shipped as a static file in the repo; this function
    # ensures it exists during `init`.  The content lives in the repo
    # itself (01_protocol/protocol_v1.0.md) and is only generated if
    # the file is missing.
    if config.PROTOCOL_MD.exists() and not force:
        log.info("Skipping (exists): %s", config.PROTOCOL_MD)
        return
    # If somehow deleted, re-create a minimal stub.
    safe_write_text(
        config.PROTOCOL_MD,
        "# SLR Protocol v1.0\n\nSee README.md for full details.\n",
        force=force,
    )


def create_amendments_log(*, force: bool = False) -> None:
    """Ensure amendments_log.csv exists."""
    safe_write_text(
        config.AMENDMENTS_CSV,
        "date,version,section,change_description,author\n",
        force=force,
    )


def create_codebook_md(*, force: bool = False) -> None:
    """Ensure codebook.md exists."""
    if config.CODEBOOK_MD.exists() and not force:
        log.info("Skipping (exists): %s", config.CODEBOOK_MD)
        return
    # File is shipped in the repo; stub if missing.
    safe_write_text(
        config.CODEBOOK_MD,
        "# Extraction Codebook\n\nSee extraction_template.xlsx Codebook sheet.\n",
        force=force,
    )


# ── Convenience: create all ────────────────────────────────────────────────

def create_all_templates(*, force: bool = False) -> None:
    """Create every template file (idempotent unless *force*)."""
    create_ta_decisions_template(force=force)
    create_ft_decisions_template(force=force)
    create_extraction_template(force=force)
    create_protocol(force=force)
    create_amendments_log(force=force)
    create_codebook_md(force=force)
