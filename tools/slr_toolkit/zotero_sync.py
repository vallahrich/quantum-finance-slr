"""Zotero API write-back: create collections, assign papers, manage tags.

Provides a ZoteroWriter class with methods for:
- Creating the Tier > Group collection hierarchy
- Finding existing items by DOI
- Adding items to collections
- Setting tags on items
- Building the paper_id_bridge.csv

Uses raw ``requests`` for consistency with the existing SLR HTTP patterns.
"""

from __future__ import annotations

import csv
import json
import logging
import os
import time
import uuid
from pathlib import Path

import requests

from . import config
from .utils import ensure_dir

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Load .env from parent repo (quantum-finance/.env)
# ---------------------------------------------------------------------------

def _find_env_file() -> Path | None:
    """Walk up from this file to find quantum-finance/.env."""
    d = Path(__file__).resolve().parent
    for _ in range(6):
        candidate = d / ".env"
        if candidate.is_file():
            return candidate
        d = d.parent
    return None

try:
    from dotenv import load_dotenv as _load_dotenv
    _env_path = _find_env_file()
    if _env_path:
        _load_dotenv(_env_path, override=False)
except ImportError:
    pass

API_BASE = "https://api.zotero.org"

# Zotero write API: max 1 request/second for writes, 50 objects per batch
_WRITE_DELAY = 1.0
_READ_DELAY = 0.2
_MAX_BATCH = 50

# Collection display names for tier hierarchy
_TIER_DISPLAY_NAMES = {
    "tier-1": "Tier 1 — General Overview",
    "tier-2": "Tier 2 — Problem-Specific",
    "tier-3": "Tier 3 — Algorithm & Experiment",
}

_TIER1_GROUP_NAMES = {
    "qc-industry": "QC Industry",
    "computational-finance": "Computational Finance",
    "qc-in-finance-overview": "QC in Finance Overview",
}

_TIER2_GROUP_NAMES = {
    "portfolio-optimization": "Portfolio Optimization",
    "derivative-pricing": "Derivative Pricing",
    "risk-management": "Risk Management",
    "fraud-detection": "Fraud Detection",
    "forecasting-prediction": "Forecasting & Prediction",
    "trading-execution": "Trading & Execution",
    "insurance-actuarial": "Insurance & Actuarial",
    "credit-lending": "Credit & Lending",
    "quantum-ml-finance": "Quantum ML for Finance",
    "optimization-methods": "Optimization Methods",
    "simulation-monte-carlo": "Simulation & Monte Carlo",
    "benchmarking-advantage": "Benchmarking & Advantage",
    "asset-pricing": "Asset Pricing",
    "quantum-cryptography": "Quantum Cryptography",
    "regulatory-compliance": "Regulatory Compliance",
}

_TIER3_GROUP_NAMES = {
    "qaoa": "QAOA",
    "vqe": "VQE",
    "amplitude-estimation": "Amplitude Estimation",
    "quantum-ml": "Quantum ML",
    "quantum-walk": "Quantum Walk",
    "hhl": "HHL",
    "hybrid": "Hybrid",
    "grover": "Grover",
    "quantum-annealing": "Quantum Annealing",
    "quantum-svm": "Quantum SVM",
    "qubo": "QUBO",
    "quantum-simulation": "Quantum Simulation",
    "classical-simulation": "Classical Simulation",
    "other-gate-based": "Other Gate-Based",
}


class ZoteroWriter:
    """Zotero API client with read and write capabilities."""

    DEFAULT_GROUP_ID = "6475432"

    def __init__(self, group_id: str = "", api_key: str = ""):
        self._group_id = group_id or os.environ.get("ZOTERO_GROUP_ID", self.DEFAULT_GROUP_ID)
        self._api_key = api_key or os.environ.get("ZOTERO_API_KEY", "")
        if not self._group_id:
            raise ValueError("ZOTERO_GROUP_ID not set")
        if not self._api_key:
            raise ValueError("ZOTERO_API_KEY not set")
        self._base = f"{API_BASE}/groups/{self._group_id}"
        self._headers = {
            "Zotero-API-Key": self._api_key,
            "Zotero-API-Version": "3",
        }
        self._library_version: int | None = None

    # ------------------------------------------------------------------
    # Low-level HTTP
    # ------------------------------------------------------------------

    def _get(self, path: str, params: dict | None = None) -> requests.Response:
        url = f"{self._base}{path}"
        for attempt in range(3):
            time.sleep(_READ_DELAY)
            resp = requests.get(url, headers=self._headers, params=params, timeout=30)
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 5))
                log.warning("Rate limited (GET), waiting %d seconds", retry_after)
                time.sleep(retry_after)
                continue
            resp.raise_for_status()
            ver = resp.headers.get("Last-Modified-Version")
            if ver:
                self._library_version = int(ver)
            return resp
        raise RuntimeError(f"GET failed after 3 retries: {url}")

    def _post(self, path: str, data: list | dict) -> requests.Response:
        url = f"{self._base}{path}"
        headers = {
            **self._headers,
            "Content-Type": "application/json",
        }
        if self._library_version is not None:
            headers["If-Unmodified-Since-Version"] = str(self._library_version)
        else:
            headers["Zotero-Write-Token"] = uuid.uuid4().hex

        for attempt in range(3):
            time.sleep(_WRITE_DELAY)
            resp = requests.post(
                url, headers=headers, json=data, timeout=60,
            )
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 5))
                log.warning("Rate limited (POST), waiting %d seconds", retry_after)
                time.sleep(retry_after)
                continue
            if resp.status_code == 412:
                log.warning("Version conflict (POST), refreshing library version")
                self._refresh_library_version()
                headers["If-Unmodified-Since-Version"] = str(self._library_version)
                continue
            resp.raise_for_status()
            ver = resp.headers.get("Last-Modified-Version")
            if ver:
                self._library_version = int(ver)
            return resp
        raise RuntimeError(f"POST failed after 3 retries: {url}")

    def _patch(self, path: str, data: dict, version: int) -> requests.Response:
        url = f"{self._base}{path}"
        for attempt in range(5):
            headers = {
                **self._headers,
                "Content-Type": "application/json",
                "If-Unmodified-Since-Version": str(version),
            }
            time.sleep(_WRITE_DELAY)
            resp = requests.patch(
                url, headers=headers, json=data, timeout=60,
            )
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 5))
                log.warning("Rate limited (PATCH), waiting %d seconds", retry_after)
                time.sleep(retry_after)
                continue
            if resp.status_code == 412:
                # Version conflict — re-fetch item to get current version
                log.warning("Version conflict (412) on %s, re-fetching...", path)
                time.sleep(_WRITE_DELAY)
                item_resp = requests.get(url, headers=self._headers, timeout=30)
                if item_resp.ok:
                    item_data = item_resp.json().get("data", item_resp.json())
                    version = item_data.get("version", version + 1)
                continue
            resp.raise_for_status()
            ver = resp.headers.get("Last-Modified-Version")
            if ver:
                self._library_version = int(ver)
            return resp
        raise RuntimeError(f"PATCH failed after 5 retries: {url}")

    def _delete(self, path: str, version: int) -> requests.Response:
        url = f"{self._base}{path}"
        headers = {
            **self._headers,
            "If-Unmodified-Since-Version": str(version),
        }
        for attempt in range(3):
            time.sleep(_WRITE_DELAY)
            resp = requests.delete(url, headers=headers, timeout=30)
            if resp.status_code == 429:
                retry_after = int(resp.headers.get("Retry-After", 5))
                log.warning("Rate limited (DELETE), waiting %d seconds", retry_after)
                time.sleep(retry_after)
                continue
            resp.raise_for_status()
            ver = resp.headers.get("Last-Modified-Version")
            if ver:
                self._library_version = int(ver)
            return resp
        raise RuntimeError(f"DELETE failed after 3 retries: {url}")

    def _refresh_library_version(self) -> None:
        resp = self._get("/collections", params={"limit": 1})
        # Library version updated in _get

    # ------------------------------------------------------------------
    # Collection operations
    # ------------------------------------------------------------------

    def list_collections(self) -> list[dict]:
        """List all collections in the group library."""
        collections = []
        start = 0
        while True:
            resp = self._get("/collections", params={
                "format": "json", "limit": 100, "start": start,
            })
            page = resp.json()
            collections.extend(page)
            total = int(resp.headers.get("Total-Results", 0))
            start += len(page)
            if start >= total or not page:
                break
        return collections

    def create_collection(self, name: str, parent_key: str | None = None) -> str:
        """Create a collection. Returns the collection key."""
        payload: dict = {"name": name}
        if parent_key:
            payload["parentCollection"] = parent_key

        resp = self._post("/collections", [payload])
        result = resp.json()

        successful = result.get("successful", result.get("success", {}))
        if "0" in successful:
            obj = successful["0"]
            key = obj["key"] if isinstance(obj, dict) else obj
            log.info("Created collection '%s' -> %s", name, key)
            return key

        failed = result.get("failed", {})
        if failed:
            msg = json.dumps(failed, indent=2)
            raise RuntimeError(f"Failed to create collection '{name}': {msg}")

        raise RuntimeError(f"Unexpected response creating collection '{name}': {result}")

    def delete_collection(self, key: str) -> bool:
        """Delete a collection by key. Returns True on success."""
        # Get current version
        resp = self._get(f"/collections/{key}")
        version = resp.json()["version"]
        self._delete(f"/collections/{key}", version=version)
        log.info("Deleted collection %s", key)
        return True

    def create_tier_hierarchy(self) -> dict:
        """Create the full Tier > Group collection hierarchy.

        Returns a mapping compatible with shared/zotero_collection_map.json.
        """
        result = {
            "version": "1.0",
            "description": "Auto-generated Zotero collection mapping",
            "generated_at": time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime()),
            "zotero_group_id": self._group_id,
            "root_collection_key": None,
            "tiers": {},
        }

        # Create root collection
        root_key = self.create_collection("Quantum Finance SLR")
        result["root_collection_key"] = root_key

        # Tier 1
        t1_key = self.create_collection(
            _TIER_DISPLAY_NAMES["tier-1"], parent_key=root_key,
        )
        t1_groups = {}
        for group_id, display_name in _TIER1_GROUP_NAMES.items():
            gkey = self.create_collection(display_name, parent_key=t1_key)
            t1_groups[group_id] = gkey
        result["tiers"]["tier-1"] = {
            "collection_key": t1_key,
            "groups": t1_groups,
        }

        # Tier 2
        t2_key = self.create_collection(
            _TIER_DISPLAY_NAMES["tier-2"], parent_key=root_key,
        )
        t2_groups = {}
        for group_id, display_name in _TIER2_GROUP_NAMES.items():
            gkey = self.create_collection(display_name, parent_key=t2_key)
            t2_groups[group_id] = gkey
        result["tiers"]["tier-2"] = {
            "collection_key": t2_key,
            "groups": t2_groups,
        }

        # Tier 3
        t3_key = self.create_collection(
            _TIER_DISPLAY_NAMES["tier-3"], parent_key=root_key,
        )
        t3_groups = {}
        for group_id, display_name in _TIER3_GROUP_NAMES.items():
            gkey = self.create_collection(display_name, parent_key=t3_key)
            t3_groups[group_id] = gkey
        result["tiers"]["tier-3"] = {
            "collection_key": t3_key,
            "groups": t3_groups,
        }

        return result

    # ------------------------------------------------------------------
    # Item operations
    # ------------------------------------------------------------------

    def get_all_items(self, limit: int = 0) -> list[dict]:
        """Fetch all non-attachment items from the library."""
        items = []
        start = 0
        page_size = min(100, limit) if limit > 0 else 100
        while True:
            resp = self._get("/items", params={
                "format": "json",
                "itemType": "-attachment || note",
                "limit": page_size,
                "start": start,
            })
            page = resp.json()
            items.extend(page)
            total = int(resp.headers.get("Total-Results", 0))
            start += len(page)
            if limit > 0 and len(items) >= limit:
                items = items[:limit]
                break
            if start >= total or not page:
                break
        return items

    @staticmethod
    def _normalize_doi(doi: str) -> str:
        """Strip URL prefixes and normalize a DOI for comparison."""
        d = doi.strip().lower()
        for prefix in ("https://doi.org/", "http://doi.org/", "https://dx.doi.org/", "http://dx.doi.org/"):
            if d.startswith(prefix):
                d = d[len(prefix):]
                break
        return d

    def find_item_by_doi(self, doi: str) -> dict | None:
        """Search for an item by DOI. Returns the first match or None."""
        if not doi:
            return None
        norm_doi = self._normalize_doi(doi)
        resp = self._get("/items", params={
            "format": "json",
            "itemType": "-attachment || note",
            "q": norm_doi,
            "qmode": "everything",
            "limit": 5,
        })
        for item in resp.json():
            item_doi = item.get("data", {}).get("DOI", "")
            if item_doi and self._normalize_doi(item_doi) == norm_doi:
                return item
        return None

    def get_item(self, item_key: str) -> dict:
        """Fetch a single item by key."""
        resp = self._get(f"/items/{item_key}")
        return resp.json()

    def add_item_to_collections(
        self, item_key: str, collection_keys: list[str],
    ) -> bool:
        """Add an item to one or more collections (without removing existing)."""
        item = self.get_item(item_key)
        data = item.get("data", item)
        version = data.get("version", item.get("version"))
        existing_collections = set(data.get("collections", []))
        new_collections = sorted(existing_collections | set(collection_keys))

        if set(new_collections) == existing_collections:
            log.debug("Item %s already in all requested collections", item_key)
            return True

        self._patch(
            f"/items/{item_key}",
            data={"collections": new_collections},
            version=version,
        )
        log.info("Added item %s to %d collections", item_key, len(collection_keys))
        return True

    def remove_item_from_collection(
        self, item_key: str, collection_key: str,
    ) -> bool:
        """Remove an item from a collection."""
        item = self.get_item(item_key)
        data = item.get("data", item)
        version = data.get("version", item.get("version"))
        collections = [c for c in data.get("collections", []) if c != collection_key]

        self._patch(
            f"/items/{item_key}",
            data={"collections": collections},
            version=version,
        )
        log.info("Removed item %s from collection %s", item_key, collection_key)
        return True

    def set_tags(self, item_key: str, tags: list[str]) -> bool:
        """Set tags on an item (replaces existing tags)."""
        item = self.get_item(item_key)
        data = item.get("data", item)
        version = data.get("version", item.get("version"))

        tag_objects = [{"tag": t} for t in tags]
        self._patch(
            f"/items/{item_key}",
            data={"tags": tag_objects},
            version=version,
        )
        log.info("Set %d tags on item %s", len(tags), item_key)
        return True

    def add_tags(self, item_key: str, new_tags: list[str]) -> bool:
        """Add tags to an item (preserving existing tags)."""
        item = self.get_item(item_key)
        data = item.get("data", item)
        version = data.get("version", item.get("version"))

        existing_tags = {t["tag"] for t in data.get("tags", [])}
        merged = sorted(existing_tags | set(new_tags))
        tag_objects = [{"tag": t} for t in merged]

        if set(merged) == existing_tags:
            log.debug("Item %s already has all requested tags", item_key)
            return True

        self._patch(
            f"/items/{item_key}",
            data={"tags": tag_objects},
            version=version,
        )
        log.info("Added %d new tags to item %s", len(set(new_tags) - existing_tags), item_key)
        return True

    def delete_item(self, item_key: str) -> bool:
        """Delete an item by key."""
        item = self.get_item(item_key)
        data = item.get("data", item)
        version = data.get("version", item.get("version"))
        self._delete(f"/items/{item_key}", version=version)
        log.info("Deleted item %s", item_key)
        return True

    # ------------------------------------------------------------------
    # SLR Results Sync
    # ------------------------------------------------------------------

    _SLR_TAGS = ["quantum-finance", "slr", "slr-results", "included"]

    @staticmethod
    def normalize_title(title: str) -> str:
        """Normalize a title for dedup comparison."""
        import re as _re
        t = title.lower().strip()
        t = _re.sub(r"[^a-z0-9]+", " ", t)
        return " ".join(t.split())

    @staticmethod
    def _extract_arxiv_id(doi: str) -> str:
        """Extract arXiv ID from DOI if present."""
        import re as _re
        if doi:
            m = _re.match(r"10\.48550/arXiv\.(.+)", doi, _re.IGNORECASE)
            if m:
                return m.group(1)
        return ""

    @staticmethod
    def _first_author_year(authors: str, year: str) -> str:
        """Build a first-author+year key for fallback dedup."""
        if not authors or not year:
            return ""
        first = authors.split(";")[0].split(",")[0].strip().lower()
        return f"{first}|{year}"

    def _build_zotero_index(self) -> tuple[
        dict[str, dict],   # doi -> item
        dict[str, dict],   # arxiv_id -> item
        dict[str, dict],   # norm_title -> item
        dict[str, dict],   # first_author_year -> item
    ]:
        """Fetch all Zotero items and build dedup indices."""
        import re as _re
        doi_idx: dict[str, dict] = {}
        arxiv_idx: dict[str, dict] = {}
        title_idx: dict[str, dict] = {}
        author_year_idx: dict[str, dict] = {}

        all_items = self.get_all_items()
        for item in all_items:
            data = item.get("data", {})
            # DOI index
            zdoi = data.get("DOI", "").strip()
            if zdoi:
                doi_idx[self._normalize_doi(zdoi)] = item
                # Also index arXiv DOIs
                aid = self._extract_arxiv_id(zdoi)
                if aid:
                    arxiv_idx[aid.lower()] = item

            # arXiv ID from Extra field (common Zotero pattern)
            extra = data.get("extra", "")
            for line in extra.split("\n"):
                line = line.strip()
                if line.lower().startswith("arxiv:"):
                    aid = line.split(":", 1)[1].strip()
                    if aid:
                        arxiv_idx[aid.lower()] = item

            # Title index
            title = data.get("title", "")
            if title:
                title_idx[self.normalize_title(title)] = item

            # First author + year
            creators = data.get("creators", [])
            if creators:
                first_creator = creators[0]
                last_name = first_creator.get("lastName", first_creator.get("name", ""))
                year = data.get("date", "")[:4]
                if last_name and year:
                    key = f"{last_name.lower().strip()}|{year}"
                    author_year_idx[key] = item

        log.info(
            "Zotero index: %d DOIs, %d arXiv IDs, %d titles, %d author-year",
            len(doi_idx), len(arxiv_idx), len(title_idx), len(author_year_idx),
        )
        return doi_idx, arxiv_idx, title_idx, author_year_idx

    def _find_existing_item(
        self,
        paper: dict,
        doi_idx: dict[str, dict],
        arxiv_idx: dict[str, dict],
        title_idx: dict[str, dict],
        author_year_idx: dict[str, dict],
    ) -> tuple[dict | None, str]:
        """Find an existing Zotero item matching this paper. Returns (item, match_method)."""
        doi = paper.get("doi", "").strip()
        if doi:
            norm_doi = self._normalize_doi(doi)
            if norm_doi in doi_idx:
                return doi_idx[norm_doi], "doi"
            # Check arXiv ID from DOI
            aid = self._extract_arxiv_id(doi)
            if aid and aid.lower() in arxiv_idx:
                return arxiv_idx[aid.lower()], "arxiv_id"

        # Title match
        title = paper.get("title", "").strip()
        if title:
            norm_t = self.normalize_title(title)
            if norm_t in title_idx:
                return title_idx[norm_t], "title"

        # First author + year fallback
        authors = paper.get("authors", "")
        year = paper.get("year", "")
        ay_key = self._first_author_year(authors, year)
        if ay_key and ay_key in author_year_idx:
            return author_year_idx[ay_key], "author_year"

        return None, ""

    @staticmethod
    def _map_paper_to_zotero_data(paper: dict) -> dict:
        """Map SLR paper metadata to Zotero item data."""
        doi = paper.get("doi", "").strip()

        # Determine item type
        source_db = paper.get("source_db", "").lower()
        venue = paper.get("venue", "").lower()
        is_preprint = paper.get("is_preprint", "") == "1"

        if is_preprint or source_db == "arxiv" or "arxiv" in venue:
            item_type = "preprint"
        elif any(kw in venue for kw in ("conference", "proceedings", "workshop", "symposium")):
            item_type = "conferencePaper"
        elif any(kw in venue for kw in ("book", "chapter", "springer", "lecture notes")):
            item_type = "bookSection"
        else:
            item_type = "journalArticle"

        # Parse authors
        creators = []
        for author_str in paper.get("authors", "").split(";"):
            author_str = author_str.strip()
            if not author_str:
                continue
            parts = author_str.split(",", 1)
            if len(parts) == 2:
                creators.append({
                    "creatorType": "author",
                    "lastName": parts[0].strip(),
                    "firstName": parts[1].strip(),
                })
            else:
                creators.append({
                    "creatorType": "author",
                    "name": author_str,
                })

        data: dict = {
            "itemType": item_type,
            "title": paper.get("title", ""),
            "creators": creators,
            "date": paper.get("year", ""),
            "DOI": doi,
            "abstractNote": paper.get("abstract", ""),
        }

        # Set venue field based on item type
        venue_val = paper.get("venue", "")
        if item_type == "journalArticle":
            data["publicationTitle"] = venue_val
        elif item_type == "conferencePaper":
            data["proceedingsTitle"] = venue_val
        elif item_type == "preprint":
            data["repository"] = venue_val or "arXiv"
        elif item_type == "bookSection":
            data["bookTitle"] = venue_val

        # Keywords
        keywords = paper.get("keywords", "").strip()
        if keywords:
            keyword_list = [k.strip() for k in keywords.split(";") if k.strip()]
            data["tags"] = [{"tag": k} for k in keyword_list]

        # Extra field for arXiv ID
        arxiv_id = ZoteroWriter._extract_arxiv_id(doi)
        if arxiv_id:
            data["extra"] = f"arXiv: {arxiv_id}"
            if item_type == "preprint":
                data["url"] = f"https://arxiv.org/abs/{arxiv_id}"

        return data

    def _create_slr_note(self, paper: dict, match_method: str = "new") -> dict:
        """Build a child note with SLR metadata."""
        from datetime import datetime as _dt

        pid = paper.get("paper_id", "")
        lines = [
            "<h2>SLR Metadata</h2>",
            f"<p><b>Paper ID:</b> {pid}</p>",
            f"<p><b>Source DB:</b> {paper.get('source_db', '')}</p>",
            f"<p><b>Inclusion reason:</b> Included in final SLR results</p>",
        ]

        # Add topic coding info if available
        if paper.get("primary_topics"):
            lines.append(f"<p><b>Primary topics:</b> {paper.get('primary_topics', '')}</p>")
        if paper.get("method_family"):
            lines.append(f"<p><b>Method family:</b> {paper.get('method_family', '')}</p>")
        if paper.get("application_area"):
            lines.append(f"<p><b>Application area:</b> {paper.get('application_area', '')}</p>")

        lines.append(f"<p><b>Sync method:</b> {match_method}</p>")
        lines.append(f"<p><b>Sync timestamp:</b> {_dt.now().isoformat(timespec='seconds')}</p>")
        lines.append(f"<p><em>Auto-synced by quantum-finance-slr toolkit</em></p>")

        return {
            "itemType": "note",
            "note": "\n".join(lines),
            "tags": [{"tag": "slr-metadata"}],
        }

    def ensure_collection(self, name: str, parent_key: str | None = None) -> str:
        """Find or create a collection by name. Returns the collection key."""
        collections = self.list_collections()
        for col in collections:
            data = col.get("data", col)
            if data.get("name") == name:
                parent = data.get("parentCollection", False)
                if parent_key is None and not parent:
                    return col["key"]
                if parent_key and parent == parent_key:
                    return col["key"]
        return self.create_collection(name, parent_key=parent_key)

    def sync_slr_results(
        self,
        *,
        collection_name: str = "SLR Results",
        dry_run: bool = True,
        max_items: int | None = None,
    ) -> dict:
        """Sync final included SLR papers into a Zotero collection.

        Parameters
        ----------
        collection_name : str
            Name of the top-level Zotero collection for results.
        dry_run : bool
            If True (default), only report what would be done.
        max_items : int | None
            Cap on papers to process (useful for testing).

        Returns
        -------
        dict
            Sync report with created/updated/skipped/failed counts and details.
        """
        report: dict = {
            "dry_run": dry_run,
            "created": [],
            "updated": [],
            "skipped": [],
            "failed": [],
        }

        # 1. Load included paper IDs
        included_ids: set[str] = set()
        with open(config.INCLUDED_FOR_CODING, encoding="utf-8", newline="") as f:
            for row in csv.DictReader(f):
                if row.get("final_decision", "").strip().lower() == "include":
                    included_ids.add(row["paper_id"])

        # 2. Load master records metadata
        master: dict[str, dict] = {}
        with open(config.MASTER_RECORDS_CSV, encoding="utf-8", newline="") as f:
            for row in csv.DictReader(f):
                if row["paper_id"] in included_ids:
                    master[row["paper_id"]] = row

        # 3. Enrich with topic coding if available
        if config.TOPIC_CODING_CSV.exists():
            with open(config.TOPIC_CODING_CSV, encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    pid = row.get("paper_id", "")
                    if pid in master:
                        for field in ("primary_topics", "secondary_topics",
                                      "application_area", "method_family"):
                            if row.get(field):
                                master[pid][field] = row[field]

        papers = [master[pid] for pid in sorted(included_ids) if pid in master]
        if max_items is not None:
            papers = papers[:max_items]

        log.info("Syncing %d papers to Zotero collection '%s'", len(papers), collection_name)

        # 4. Build dedup index from existing Zotero library
        print(f"Building Zotero library index...")
        doi_idx, arxiv_idx, title_idx, ay_idx = self._build_zotero_index()

        # 5. Find or create the target collection
        collection_key = None
        if not dry_run:
            collection_key = self.ensure_collection(collection_name)
            log.info("Target collection '%s' -> %s", collection_name, collection_key)
        print(f"Target collection: '{collection_name}'" +
              (f" [{collection_key}]" if collection_key else " [dry-run]"))

        # 6. Process each paper
        for i, paper in enumerate(papers, 1):
            pid = paper["paper_id"]
            title = paper.get("title", "") or pid
            safe_title = title[:60] + "..." if len(title) > 60 else title
            safe_title = safe_title.encode("ascii", errors="replace").decode("ascii")

            existing, match_method = self._find_existing_item(
                paper, doi_idx, arxiv_idx, title_idx, ay_idx,
            )

            if existing:
                zkey = existing["key"]
                print(f"  [{i}/{len(papers)}] UPDATE ({match_method}): {safe_title}")

                if dry_run:
                    report["updated"].append({
                        "paper_id": pid, "zotero_key": zkey,
                        "match_method": match_method, "title": title,
                    })
                    continue

                try:
                    # Add to collection + add SLR tags
                    item_data = existing.get("data", existing)
                    version = item_data.get("version", existing.get("version"))

                    patch: dict = {}
                    # Merge collections
                    existing_cols = set(item_data.get("collections", []))
                    if collection_key and collection_key not in existing_cols:
                        patch["collections"] = sorted(existing_cols | {collection_key})

                    # Merge tags
                    existing_tags = {t["tag"] for t in item_data.get("tags", [])}
                    new_tags = [t for t in self._SLR_TAGS if t not in existing_tags]
                    if new_tags:
                        all_tags = item_data.get("tags", []) + [{"tag": t} for t in new_tags]
                        patch["tags"] = all_tags

                    if patch:
                        self._patch(f"/items/{zkey}", data=patch, version=version)

                    # Add child note
                    note_data = self._create_slr_note(paper, match_method=match_method)
                    note_data["parentItem"] = zkey
                    if collection_key:
                        note_data["collections"] = [collection_key]
                    self._post("/items", [note_data])

                    report["updated"].append({
                        "paper_id": pid, "zotero_key": zkey,
                        "match_method": match_method, "title": title,
                    })
                except Exception as exc:
                    log.warning("Failed to update %s: %s", pid, exc)
                    report["failed"].append({
                        "paper_id": pid, "error": str(exc), "title": title,
                    })
            else:
                print(f"  [{i}/{len(papers)}] CREATE: {safe_title}")

                if dry_run:
                    report["created"].append({
                        "paper_id": pid, "title": title,
                    })
                    continue

                try:
                    item_data = self._map_paper_to_zotero_data(paper)

                    # Add SLR tags
                    existing_tags = item_data.get("tags", [])
                    for t in self._SLR_TAGS:
                        if not any(tag["tag"] == t for tag in existing_tags):
                            existing_tags.append({"tag": t})
                    item_data["tags"] = existing_tags

                    # Add to collection
                    if collection_key:
                        item_data["collections"] = [collection_key]

                    resp = self._post("/items", [item_data])
                    result = resp.json()
                    successful = result.get("successful", result.get("success", {}))
                    if "0" in successful:
                        obj = successful["0"]
                        new_key = obj["key"] if isinstance(obj, dict) else obj
                        log.info("Created item %s for paper %s", new_key, pid)

                        # Add child note
                        note_data = self._create_slr_note(paper, match_method="new")
                        note_data["parentItem"] = new_key
                        if collection_key:
                            note_data["collections"] = [collection_key]
                        self._post("/items", [note_data])

                        report["created"].append({
                            "paper_id": pid, "zotero_key": new_key, "title": title,
                        })
                    else:
                        failed = result.get("failed", {})
                        raise RuntimeError(f"Zotero create failed: {failed}")
                except Exception as exc:
                    log.warning("Failed to create %s: %s", pid, exc)
                    report["failed"].append({
                        "paper_id": pid, "error": str(exc), "title": title,
                    })

        # 7. Summary
        report["summary"] = {
            "total": len(papers),
            "created": len(report["created"]),
            "updated": len(report["updated"]),
            "skipped": len(report["skipped"]),
            "failed": len(report["failed"]),
        }
        return report

    # ------------------------------------------------------------------
    # Bridge building
    # ------------------------------------------------------------------

    @staticmethod
    def _load_dois_from_reviewers(
        paper_ids: set[str],
    ) -> tuple[dict[str, str], dict[str, str]]:
        """Load DOIs and titles from reviewer spreadsheets."""
        try:
            import openpyxl
        except ImportError:
            log.warning("openpyxl not installed, cannot read reviewer spreadsheets")
            return {}, {}

        doi_map: dict[str, str] = {}
        title_map: dict[str, str] = {}
        screening_dir = config.ROOT_DIR / "05_screening"

        for fname, sheet in [
            ("screening_reviewer_A.xlsx", None),
            ("screening_reviewer_B.xlsx", "Screening"),
        ]:
            fpath = screening_dir / fname
            if not fpath.exists():
                continue
            wb = openpyxl.load_workbook(str(fpath), read_only=True)
            ws = wb[sheet] if sheet else wb.active
            headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if "Paper ID" not in headers or "DOI" not in headers:
                wb.close()
                continue
            pidx = headers.index("Paper ID")
            didx = headers.index("DOI")
            tidx = headers.index("Title") if "Title" in headers else -1
            for row in ws.iter_rows(min_row=2, values_only=True):
                pid = row[pidx]
                if not pid or pid not in paper_ids:
                    continue
                doi = str(row[didx] or "").strip()
                if doi and pid not in doi_map:
                    doi_map[pid] = doi
                if tidx >= 0 and pid not in title_map:
                    title_map[pid] = str(row[tidx] or "")
            wb.close()
            log.info("Loaded %d DOIs from %s", len(doi_map), fname)

        return doi_map, title_map

    def build_paper_id_bridge(
        self,
        topic_csv_path: Path | None = None,
        master_csv_path: Path | None = None,
        bridge_output_path: Path | None = None,
    ) -> int:
        """Match SLR papers to Zotero items by DOI. Returns count of matches.

        Reads DOIs from master_records.csv or reviewer spreadsheets.
        Queries Zotero for each DOI. Writes bridge CSV.
        """
        if topic_csv_path is None:
            topic_csv_path = config.TOPIC_CODING_CSV
        if master_csv_path is None:
            master_csv_path = config.MASTER_RECORDS_CSV

        # Collect paper_ids from topic coding
        with open(topic_csv_path, encoding="utf-8", newline="") as f:
            paper_ids = {
                row["paper_id"]
                for row in csv.DictReader(f)
                if row.get("paper_id")
            }

        # Load DOIs from master records
        doi_map: dict[str, str] = {}  # paper_id -> doi
        title_map: dict[str, str] = {}  # paper_id -> title
        if master_csv_path.exists():
            with open(master_csv_path, encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    pid = row.get("paper_id", "")
                    if pid in paper_ids:
                        doi = row.get("doi", "").strip()
                        if doi:
                            doi_map[pid] = doi
                        title_map[pid] = row.get("title", "")

        # Fallback: load DOIs from reviewer spreadsheets
        if not doi_map:
            log.info("master_records.csv unavailable, loading DOIs from reviewer spreadsheets")
            doi_map, title_map = self._load_dois_from_reviewers(paper_ids)

        # Also load DOIs from existing bridge file (may have been enriched externally)
        if bridge_output_path is None:
            _bridge_path = config.ROOT_DIR.parent / "shared" / "paper_id_bridge.csv"
        else:
            _bridge_path = bridge_output_path
        if _bridge_path.exists():
            with open(_bridge_path, encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    pid = row.get("slr_paper_id", "")
                    doi = row.get("doi", "").strip()
                    if pid in paper_ids and doi and pid not in doi_map:
                        doi_map[pid] = doi
                        if not title_map.get(pid):
                            title_map[pid] = row.get("title", "")
            log.info("After merging existing bridge: %d DOIs total", len(doi_map))

        if not doi_map:
            log.warning(
                "No DOIs found for topic-coded papers. "
                "Bridge building requires DOIs from master_records.csv or reviewer spreadsheets."
            )
            return 0

        # Fetch ALL Zotero items once and build a DOI->key lookup
        log.info("Fetching all Zotero items for bulk DOI matching...")
        zotero_doi_index: dict[str, dict] = {}  # normalized_doi -> item
        start = 0
        page_size = 100
        while True:
            resp = self._get("/items", params={
                "format": "json",
                "itemType": "-attachment || note",
                "limit": page_size,
                "start": start,
            })
            items = resp.json()
            if not items:
                break
            for item in items:
                zdoi = item.get("data", {}).get("DOI", "")
                if zdoi:
                    norm = self._normalize_doi(zdoi)
                    zotero_doi_index[norm] = item
            start += len(items)
            if len(items) < page_size:
                break
        log.info("Indexed %d Zotero items with DOIs", len(zotero_doi_index))

        # Match
        bridge_rows: list[dict[str, str]] = []
        matched = 0

        for pid, doi in sorted(doi_map.items()):
            norm_doi = self._normalize_doi(doi)
            item = zotero_doi_index.get(norm_doi)
            if item:
                bridge_rows.append({
                    "slr_paper_id": pid,
                    "doi": doi,
                    "zotero_item_key": item["key"],
                    "title": title_map.get(pid, ""),
                    "match_method": "doi",
                })
                matched += 1
            else:
                bridge_rows.append({
                    "slr_paper_id": pid,
                    "doi": doi,
                    "zotero_item_key": "",
                    "title": title_map.get(pid, ""),
                    "match_method": "unmatched",
                })

        # Write bridge CSV
        if bridge_output_path is None:
            # Default to parent repo shared/ directory if possible
            bridge_output_path = config.ROOT_DIR.parent / "shared" / "paper_id_bridge.csv"
        ensure_dir(bridge_output_path.parent)

        with open(bridge_output_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "slr_paper_id", "doi", "zotero_item_key", "title", "match_method",
            ])
            writer.writeheader()
            writer.writerows(bridge_rows)

        log.info(
            "Bridge built: %d matched, %d unmatched, %d total",
            matched, len(bridge_rows) - matched, len(bridge_rows),
        )
        return matched

    # ------------------------------------------------------------------
    # Tier assignment
    # ------------------------------------------------------------------

    def assign_papers_to_tiers(
        self,
        tier_csv_path: Path | None = None,
        bridge_csv_path: Path | None = None,
        collection_map: dict | None = None,
    ) -> dict[str, int]:
        """Assign papers to Zotero collections and tags based on tier classification.

        Only processes papers with review_status == 'approved'.

        Returns dict with counts: assigned, skipped_not_approved, skipped_no_zotero.
        """
        if tier_csv_path is None:
            tier_csv_path = config.TIER_CLASSIFICATION_CSV

        if not tier_csv_path.exists():
            raise FileNotFoundError(
                f"Tier classification file not found: {tier_csv_path}. "
                "Run 'classify-tiers' first."
            )

        # Load bridge for paper_id -> zotero_item_key mapping
        zotero_map: dict[str, str] = {}  # paper_id -> zotero_item_key
        if bridge_csv_path:
            bridge_path = bridge_csv_path
        else:
            bridge_path = config.ROOT_DIR.parent / "shared" / "paper_id_bridge.csv"

        if bridge_path.exists():
            with open(bridge_path, encoding="utf-8", newline="") as f:
                for row in csv.DictReader(f):
                    pid = row.get("slr_paper_id", "")
                    zkey = row.get("zotero_item_key", "")
                    if pid and zkey:
                        zotero_map[pid] = zkey
        else:
            log.warning("Bridge file not found at %s", bridge_path)

        # Load tier classifications
        with open(tier_csv_path, encoding="utf-8", newline="") as f:
            tier_rows = list(csv.DictReader(f))

        counts = {"assigned": 0, "skipped_not_approved": 0, "skipped_no_zotero": 0}

        for row in tier_rows:
            pid = row.get("paper_id", "")
            status = row.get("review_status", "").strip().lower()

            if status != "approved":
                counts["skipped_not_approved"] += 1
                continue

            zotero_key = zotero_map.get(pid)
            if not zotero_key:
                log.warning("No Zotero item for paper %s — skipping", pid)
                counts["skipped_no_zotero"] += 1
                continue

            # Determine collections to add
            collection_keys = []
            tags = []

            if collection_map:
                tiers_data = collection_map.get("tiers", {})

                # Tier 1
                t1_groups = _read_json_array(row.get("tier1_groups", ""))
                t1_data = tiers_data.get("tier-1", {}).get("groups", {})
                for g in t1_groups:
                    ckey = t1_data.get(g)
                    if ckey:
                        collection_keys.append(ckey)
                    tags.append(f"tier:1-{g}")

                # Tier 2
                t2_groups = _read_json_array(row.get("tier2_groups", ""))
                t2_data = tiers_data.get("tier-2", {}).get("groups", {})
                for g in t2_groups:
                    ckey = t2_data.get(g)
                    if ckey:
                        collection_keys.append(ckey)
                    tags.append(f"tier:2-{g}")

                # Tier 3
                t3_groups = _read_json_array(row.get("tier3_groups", ""))
                t3_data = tiers_data.get("tier-3", {}).get("groups", {})
                for g in t3_groups:
                    ckey = t3_data.get(g)
                    if ckey:
                        collection_keys.append(ckey)
                    tags.append(f"tier:3-{g}")

            if collection_keys or tags:
                # Single PATCH: fetch item, merge collections + tags, write once
                item = self.get_item(zotero_key)
                data = item.get("data", item)
                version = data.get("version", item.get("version"))
                patch_data: dict = {}

                if collection_keys:
                    existing_cols = set(data.get("collections", []))
                    merged_cols = sorted(existing_cols | set(collection_keys))
                    if set(merged_cols) != existing_cols:
                        patch_data["collections"] = merged_cols

                if tags:
                    existing_tags = {t["tag"] for t in data.get("tags", [])}
                    new_tags = [t for t in tags if t not in existing_tags]
                    if new_tags:
                        all_tags = data.get("tags", []) + [{"tag": t} for t in new_tags]
                        patch_data["tags"] = all_tags

                if patch_data:
                    self._patch(f"/items/{zotero_key}", data=patch_data, version=version)

            counts["assigned"] += 1
            log.info(
                "Assigned paper %s (Zotero: %s) to %d collections with %d tags",
                pid, zotero_key, len(collection_keys), len(tags),
            )

        return counts


def _read_json_array(raw: str) -> list[str]:
    """Parse a JSON array string, returning empty list on failure."""
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError:
        return []
    if not isinstance(parsed, list):
        return []
    return [str(item) for item in parsed if str(item).strip()]
