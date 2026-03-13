"""Tests for AI-assisted screening pipeline (Protocol §8, Amendment A8)."""

from __future__ import annotations

import csv
import textwrap
from pathlib import Path

import pytest

from tools.slr_toolkit import config


# ── Fixtures ──────────────────────────────────────────────────────────────


def _make_master_csv(path: Path, records: list[dict]) -> Path:
    """Write a minimal master_records.csv."""
    cols = config.NORMALIZED_COLUMNS + ["duplicate_of"]
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=cols)
        writer.writeheader()
        for rec in records:
            row = {c: rec.get(c, "") for c in cols}
            writer.writerow(row)
    return path


def _make_calibration_xlsx(path: Path, decisions: list[tuple[str, str, str, str]]) -> Path:
    """Write a minimal calibration screening Excel.

    Each tuple: (paper_id, title, reviewer_a_decision, final_decision).
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Screening"
    headers = [
        "#", "Paper ID", "Title", "Authors", "Year", "DOI",
        "Abstract", "Source", "Reviewer A Decision",
        "Reviewer B Decision", "Final Decision", "Notes",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)

    for idx, (pid, title, rev_a, final) in enumerate(decisions, start=2):
        ws.cell(row=idx, column=1, value=idx - 1)
        ws.cell(row=idx, column=2, value=pid)
        ws.cell(row=idx, column=3, value=title)
        ws.cell(row=idx, column=7, value=f"Abstract of {title}")
        ws.cell(row=idx, column=9, value=rev_a)
        ws.cell(row=idx, column=11, value=final)

    wb.save(path)
    return path


SAMPLE_RECORDS = [
    {"paper_id": f"p{i:03d}", "title": f"Paper {i}", "authors": f"Author {i}",
     "year": "2023", "abstract": f"Abstract for paper {i}", "source_db": "test"}
    for i in range(20)
]


# ── Tests: export_asreview_dataset ────────────────────────────────────────


class TestExportASReviewDataset:
    def test_exports_all_unique_records(self, tmp_path, monkeypatch):
        from tools.slr_toolkit.screening import export_asreview_dataset

        master = _make_master_csv(tmp_path / "master_records.csv", SAMPLE_RECORDS)
        monkeypatch.setattr(config, "MASTER_RECORDS_CSV", master)

        out = export_asreview_dataset(output_path=tmp_path / "dataset.csv")
        assert out.exists()

        with open(out, encoding="utf-8") as f:
            reader = list(csv.DictReader(f))
        assert len(reader) == 20
        assert "title" in reader[0]
        assert "abstract" in reader[0]
        assert "paper_id" in reader[0]

    def test_excludes_specified_ids(self, tmp_path, monkeypatch):
        from tools.slr_toolkit.screening import export_asreview_dataset

        master = _make_master_csv(tmp_path / "master_records.csv", SAMPLE_RECORDS)
        monkeypatch.setattr(config, "MASTER_RECORDS_CSV", master)

        out = export_asreview_dataset(
            output_path=tmp_path / "dataset.csv",
            exclude_ids={"p000", "p001", "p002"},
        )
        with open(out, encoding="utf-8") as f:
            reader = list(csv.DictReader(f))
        assert len(reader) == 17


# ── Tests: export_asreview_labels ─────────────────────────────────────────


class TestExportASReviewLabels:
    def test_exports_consensus_labels(self, tmp_path):
        from tools.slr_toolkit.screening import export_asreview_labels

        cal = _make_calibration_xlsx(
            tmp_path / "cal.xlsx",
            [
                ("p001", "Paper 1", "include", "include"),
                ("p002", "Paper 2", "exclude", "exclude"),
                ("p003", "Paper 3", "maybe", "maybe"),  # should be skipped
            ],
        )
        out = export_asreview_labels(cal, tmp_path / "labels.csv")
        assert out.exists()

        with open(out, encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 2  # maybe excluded
        assert rows[0]["label_included"] == "1"
        assert rows[1]["label_included"] == "0"


# ── Tests: import_ai_decisions ────────────────────────────────────────────


class TestImportAIDecisions:
    def test_standard_label_included(self, tmp_path):
        from tools.slr_toolkit.screening import import_ai_decisions

        ai_csv = tmp_path / "ai_export.csv"
        ai_csv.write_text(
            "paper_id,label_included,confidence\n"
            "p001,1,0.95\n"
            "p002,0,0.30\n"
        )
        out = import_ai_decisions(ai_csv, output_path=tmp_path / "ai_decisions.csv")
        with open(out, encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 2
        assert rows[0]["ai_decision"] == "include"
        assert rows[0]["ai_confidence"] == "0.95"
        assert rows[1]["ai_decision"] == "exclude"

    def test_label_column_variant(self, tmp_path):
        from tools.slr_toolkit.screening import import_ai_decisions

        ai_csv = tmp_path / "ai_export.csv"
        ai_csv.write_text(
            "paper_id,label\n"
            "p001,relevant\n"
            "p002,irrelevant\n"
        )
        out = import_ai_decisions(ai_csv, output_path=tmp_path / "ai_decisions.csv")
        with open(out, encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["ai_decision"] == "include"
        assert rows[1]["ai_decision"] == "exclude"

    def test_missing_label_column_raises(self, tmp_path):
        from tools.slr_toolkit.screening import import_ai_decisions

        ai_csv = tmp_path / "ai_export.csv"
        ai_csv.write_text("paper_id,title\np001,X\n")
        with pytest.raises(ValueError, match="label"):
            import_ai_decisions(ai_csv, output_path=tmp_path / "out.csv")

    def test_record_id_fallback(self, tmp_path):
        from tools.slr_toolkit.screening import import_ai_decisions

        ai_csv = tmp_path / "ai_export.csv"
        ai_csv.write_text("record_id,included\np001,1\n")
        out = import_ai_decisions(ai_csv, output_path=tmp_path / "out.csv")
        with open(out, encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert rows[0]["paper_id"] == "p001"


# ── Tests: find_discrepancies ─────────────────────────────────────────────


class TestFindDiscrepancies:
    def test_classifies_discrepancy_types(self, tmp_path, monkeypatch):
        from tools.slr_toolkit.screening import find_discrepancies

        # Human decisions
        human_csv = tmp_path / "human.csv"
        human_csv.write_text(
            "paper_id,decision\n"
            "p001,include\n"
            "p002,exclude\n"
            "p003,include\n"
            "p004,exclude\n"
        )

        # AI decisions
        ai_csv = tmp_path / "ai.csv"
        ai_csv.write_text(
            "paper_id,ai_decision,ai_confidence\n"
            "p001,include,0.9\n"
            "p002,include,0.8\n"   # ai_rescue
            "p003,exclude,0.4\n"   # human_only
            "p004,exclude,0.2\n"   # agree_exclude
        )

        # Monkeypatch master for title lookup
        master = _make_master_csv(
            tmp_path / "master.csv",
            [{"paper_id": f"p{i:03d}", "title": f"T{i}"} for i in range(1, 5)],
        )
        monkeypatch.setattr(config, "MASTER_RECORDS_CSV", master)

        out_path = tmp_path / "disc.csv"
        counts = find_discrepancies(human_csv, ai_csv, out_path)

        assert counts["ai_rescue"] == 1
        assert counts["agree_include"] == 1
        assert counts["agree_exclude"] == 1
        assert counts["human_only"] == 1

        with open(out_path, encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 4
        assert any(r["discrepancy_type"] == "ai_rescue" for r in rows)


# ── Tests: generate_fn_audit ──────────────────────────────────────────────


class TestGenerateFNAudit:
    def test_samples_double_excluded(self, tmp_path):
        from tools.slr_toolkit.screening import generate_fn_audit

        disc_csv = tmp_path / "disc.csv"
        # 20 agree_exclude + 2 ai_rescue
        rows = []
        for i in range(20):
            rows.append(f"p{i:03d},T{i},exclude,exclude,0.1,agree_exclude,,")
        rows.append("p100,T100,exclude,include,0.8,ai_rescue,,")
        rows.append("p101,T101,include,include,0.9,agree_include,,")

        disc_csv.write_text(
            "paper_id,title,human_decision,ai_decision,ai_confidence,"
            "discrepancy_type,re_review_decision,notes\n"
            + "\n".join(rows) + "\n"
        )

        out = generate_fn_audit(disc_csv, tmp_path / "audit.csv", audit_fraction=0.10)
        with open(out, encoding="utf-8") as f:
            audit_rows = list(csv.DictReader(f))
        assert len(audit_rows) == 2  # 10% of 20 = 2
        assert "audit_decision" in audit_rows[0]


# ── Tests: compute_ai_validation ──────────────────────────────────────────


class TestComputeAIValidation:
    def test_perfect_recall(self, tmp_path, monkeypatch):
        from tools.slr_toolkit.screening import compute_ai_validation

        # Validation workbook: 4 records, 2 include + 2 exclude
        val_xlsx = _make_calibration_xlsx(
            tmp_path / "val.xlsx",
            [
                ("p001", "T1", "include", "include"),
                ("p002", "T2", "include", "include"),
                ("p003", "T3", "exclude", "exclude"),
                ("p004", "T4", "exclude", "exclude"),
            ],
        )

        # AI: perfect match
        ai_csv = tmp_path / "ai.csv"
        ai_csv.write_text(
            "paper_id,ai_decision,ai_confidence\n"
            "p001,include,0.99\n"
            "p002,include,0.95\n"
            "p003,exclude,0.10\n"
            "p004,exclude,0.05\n"
        )

        report = tmp_path / "report.md"
        metrics = compute_ai_validation(val_xlsx, ai_csv, report)

        assert metrics["recall"] == 1.0
        assert metrics["specificity"] == 1.0
        assert metrics["pass"] is True
        assert report.exists()
        assert "PASS" in report.read_text()

    def test_low_recall_fails(self, tmp_path):
        from tools.slr_toolkit.screening import compute_ai_validation

        val_xlsx = _make_calibration_xlsx(
            tmp_path / "val.xlsx",
            [
                ("p001", "T1", "include", "include"),
                ("p002", "T2", "include", "include"),
                ("p003", "T3", "exclude", "exclude"),
            ],
        )

        # AI misses one include → recall = 0.5
        ai_csv = tmp_path / "ai.csv"
        ai_csv.write_text(
            "paper_id,ai_decision,ai_confidence\n"
            "p001,include,0.9\n"
            "p002,exclude,0.4\n"
            "p003,exclude,0.1\n"
        )

        metrics = compute_ai_validation(val_xlsx, ai_csv, tmp_path / "report.md")
        assert metrics["recall"] == 0.5
        assert metrics["pass"] is False
