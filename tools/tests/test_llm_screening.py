"""Tests for LLM-based screening module (Protocol Amendment A9)."""

from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest

from tools.slr_toolkit import config
from .conftest import SAMPLE_RECORDS, make_master_csv


# â”€â”€ Tests: _build_url â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


class TestBuildUrl:
    def test_standard_azure_endpoint(self):
        from tools.slr_toolkit.llm_screening import _build_url

        url = _build_url("https://myresource.openai.azure.com", "gpt-4o")
        assert "deployments/gpt-4o" in url
        assert "api-version=" in url

    def test_v1_compatible_endpoint(self):
        from tools.slr_toolkit.llm_screening import _build_url

        url = _build_url(
            "https://myresource.openai.azure.com/openai/v1/", "gpt-4o",
        )
        assert url.endswith("/chat/completions")
        assert "deployments" not in url

    def test_trailing_slash_stripped(self):
        from tools.slr_toolkit.llm_screening import _build_url

        url = _build_url("https://myresource.openai.azure.com/", "gpt-4o")
        assert "deployments/gpt-4o" in url


# â”€â”€ Tests: _parse_llm_response â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


class TestParseLLMResponse:
    def test_valid_json(self):
        from tools.slr_toolkit.llm_screening import _parse_llm_response

        raw = {
            "choices": [{"message": {"content": json.dumps({
                "decision": "include",
                "confidence": 0.85,
                "reason_code": "INCLUDE",
                "reasoning": "Gate-based quantum for portfolio optimization.",
            })}}],
            "usage": {"prompt_tokens": 100, "completion_tokens": 50},
        }
        result = _parse_llm_response(raw)
        assert result["decision"] == "include"
        assert result["confidence"] == 0.85
        assert result["reason_code"] == "INCLUDE"

    def test_borderline_mapped_to_include(self):
        from tools.slr_toolkit.llm_screening import _parse_llm_response

        raw = {
            "choices": [{"message": {"content": json.dumps({
                "decision": "borderline",
                "confidence": 0.5,
                "reason_code": "INCLUDE",
                "reasoning": "Unclear scope.",
            })}}],
            "usage": {},
        }
        result = _parse_llm_response(raw)
        assert result["decision"] == "include"
        assert "borderline" in result["reasoning"]
        assert "include" in result["reasoning"]

    def test_json_in_markdown_fence(self):
        from tools.slr_toolkit.llm_screening import _parse_llm_response

        raw = {
            "choices": [{"message": {"content":
                '```json\n{"decision": "exclude", "confidence": 0.9, '
                '"reason_code": "EX-PARADIGM", "reasoning": "Annealing only."}\n```'
            }}],
            "usage": {},
        }
        result = _parse_llm_response(raw)
        assert result["decision"] == "exclude"
        assert result["reason_code"] == "EX-PARADIGM"

    def test_missing_decision_raises(self):
        from tools.slr_toolkit.llm_screening import _parse_llm_response

        raw = {
            "choices": [{"message": {"content": '{"confidence": 0.5}'}}],
            "usage": {},
        }
        with pytest.raises(ValueError, match="Missing 'decision'"):
            _parse_llm_response(raw)

    def test_no_choices_raises(self):
        from tools.slr_toolkit.llm_screening import _parse_llm_response

        with pytest.raises(ValueError, match="No choices"):
            _parse_llm_response({"choices": [], "usage": {}})

    def test_confidence_clamped(self):
        from tools.slr_toolkit.llm_screening import _parse_llm_response

        raw = {
            "choices": [{"message": {"content": json.dumps({
                "decision": "include", "confidence": 1.5,
            })}}],
            "usage": {},
        }
        result = _parse_llm_response(raw)
        assert result["confidence"] == 1.0

    def test_defaults_for_missing_fields(self):
        from tools.slr_toolkit.llm_screening import _parse_llm_response

        raw = {
            "choices": [{"message": {"content": json.dumps({
                "decision": "exclude",
            })}}],
            "usage": {},
        }
        result = _parse_llm_response(raw)
        assert result["reason_code"] == "EX-OTHER"
        assert result["confidence"] == 0.5

    def test_invalid_reason_code_falls_back_to_schema_default(self):
        from tools.slr_toolkit.llm_screening import _parse_llm_response

        raw = {
            "choices": [{"message": {"content": json.dumps({
                "decision": "exclude",
                "reason_code": "not-a-real-code",
            })}}],
            "usage": {},
        }
        result = _parse_llm_response(raw)
        assert result["reason_code"] == "EX-OTHER"


# â”€â”€ Tests: estimate_cost â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


class TestEstimateCost:
    def test_returns_cost_breakdown(self):
        from tools.slr_toolkit.llm_screening import estimate_cost

        records = [
            {"paper_id": "p001", "title": "Quantum finance", "abstract": "Text"},
            {"paper_id": "p002", "title": "Another paper", "abstract": "More text"},
        ]
        result = estimate_cost(records)
        assert result["n_records"] == 2
        assert result["est_total_tokens"] > 0
        assert result["est_total_cost_usd"] > 0

    def test_empty_records(self):
        from tools.slr_toolkit.llm_screening import estimate_cost

        result = estimate_cost([])
        assert result["n_records"] == 0
        assert result["est_total_cost_usd"] == 0


# â”€â”€ Tests: checkpoint â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


class TestCheckpoint:
    def test_save_and_load(self, tmp_path):
        from tools.slr_toolkit.llm_screening import _load_checkpoint, _save_checkpoint

        cp = tmp_path / "checkpoint.json"
        state = {"screened_ids": ["p001", "p002"], "version": 1}
        _save_checkpoint(cp, state)

        loaded = _load_checkpoint(cp)
        assert loaded["screened_ids"] == ["p001", "p002"]

    def test_load_nonexistent(self, tmp_path):
        from tools.slr_toolkit.llm_screening import _load_checkpoint

        loaded = _load_checkpoint(tmp_path / "nonexistent.json")
        assert loaded["screened_ids"] == []

    def test_load_corrupt_checkpoint_returns_default_state(self, tmp_path):
        from tools.slr_toolkit.llm_screening import _load_checkpoint

        checkpoint = tmp_path / "checkpoint.json"
        checkpoint.write_text("{not valid json", encoding="utf-8")

        loaded = _load_checkpoint(checkpoint)
        assert loaded["screened_ids"] == []
        assert loaded["version"] == 1


# â”€â”€ Tests: _write_decisions_csv â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


class TestWriteDecisions:
    def test_ranks_by_confidence_descending(self, tmp_path):
        from tools.slr_toolkit.llm_screening import _write_decisions_csv

        rows = [
            {"paper_id": "p001", "ai_decision": "include", "ai_confidence": "0.5",
             "reason_code": "INCLUDE", "reasoning": "test"},
            {"paper_id": "p002", "ai_decision": "exclude", "ai_confidence": "0.9",
             "reason_code": "EX-PARADIGM", "reasoning": "test"},
        ]
        out = tmp_path / "decisions.csv"
        _write_decisions_csv(out, rows)

        with open(out, encoding="utf-8") as f:
            result = list(csv.DictReader(f))
        assert result[0]["paper_id"] == "p002"
        assert result[0]["ai_rank"] == "1"
        assert result[1]["paper_id"] == "p001"
        assert result[1]["ai_rank"] == "2"

    def test_output_has_all_required_columns(self, tmp_path):
        from tools.slr_toolkit.llm_screening import _write_decisions_csv

        rows = [{"paper_id": "p001", "ai_decision": "include", "ai_confidence": "0.8"}]
        out = tmp_path / "decisions.csv"
        _write_decisions_csv(out, rows)

        with open(out, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            cols = reader.fieldnames
        assert "paper_id" in cols
        assert "ai_decision" in cols
        assert "ai_rank" in cols
        assert "ai_confidence" in cols


# â”€â”€ Tests: downstream compatibility â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€


class TestDownstreamCompatibility:
    """Verify the LLM output CSV is consumable by find_discrepancies and
    compute_ai_validation exactly like the ASReview output."""

    def _write_llm_decisions(self, path: Path) -> None:
        from tools.slr_toolkit.llm_screening import _write_decisions_csv
        _write_decisions_csv(path, [
            {"paper_id": "p001", "ai_decision": "include", "ai_confidence": "0.9",
             "reason_code": "INCLUDE", "reasoning": "quantum finance paper"},
            {"paper_id": "p002", "ai_decision": "include", "ai_confidence": "0.8",
             "reason_code": "INCLUDE", "reasoning": "portfolio optimization"},
            {"paper_id": "p003", "ai_decision": "exclude", "ai_confidence": "0.4",
             "reason_code": "EX-PARADIGM", "reasoning": "annealing only"},
            {"paper_id": "p004", "ai_decision": "exclude", "ai_confidence": "0.2",
             "reason_code": "EX-NONFIN", "reasoning": "chemistry application"},
        ])

    def test_find_discrepancies_accepts_llm_output(self, tmp_path, monkeypatch):
        from tools.slr_toolkit.screening import find_discrepancies

        ai_csv = tmp_path / "ai_decisions.csv"
        self._write_llm_decisions(ai_csv)

        human_csv = tmp_path / "human.csv"
        human_csv.write_text(
            "paper_id,decision\n"
            "p001,include\np002,exclude\np003,include\np004,exclude\n"
        )

        master = make_master_csv(
            tmp_path / "master.csv",
            [{"paper_id": f"p{i:03d}", "title": f"T{i}"} for i in range(1, 5)],
        )
        monkeypatch.setattr(config, "MASTER_RECORDS_CSV", master)

        counts = find_discrepancies(human_csv, ai_csv, tmp_path / "disc.csv")

        assert counts["agree_include"] == 1
        assert counts["ai_rescue"] == 1
        assert counts["human_only"] == 1
        assert counts["agree_exclude"] == 1

    def test_compute_ai_validation_accepts_llm_output(self, tmp_path):
        from tools.slr_toolkit.screening import compute_ai_validation
        from openpyxl import Workbook

        # Validation workbook with known decisions
        val_path = tmp_path / "val.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Screening"
        for i, h in enumerate(["#", "Paper ID", "Title", "Authors", "Year",
                                "DOI", "Abstract", "Source",
                                "Rev A", "Rev B", "Final", "Notes"], 1):
            ws.cell(row=1, column=i, value=h)
        for idx, (pid, dec) in enumerate([
            ("p001", "include"), ("p002", "include"),
            ("p003", "exclude"), ("p004", "exclude"),
        ], start=2):
            ws.cell(row=idx, column=1, value=idx - 1)
            ws.cell(row=idx, column=2, value=pid)
            ws.cell(row=idx, column=3, value=f"Title {pid}")
            ws.cell(row=idx, column=9, value=dec)
            ws.cell(row=idx, column=11, value=dec)
        wb.save(val_path)

        # LLM decisions CSV
        ai_csv = tmp_path / "ai.csv"
        self._write_llm_decisions(ai_csv)

        metrics = compute_ai_validation(val_path, ai_csv, tmp_path / "report.md")

        assert metrics["n"] == 4
        assert "recall" in metrics
        assert "pass" in metrics


class TestRunLLMScreeningValidation:
    def test_rejects_non_positive_batch_size(self):
        from tools.slr_toolkit.llm_screening import run_llm_screening

        with pytest.raises(ValueError, match="batch_size"):
            run_llm_screening(batch_size=0, estimate_only=True)

    def test_rejects_negative_delay(self):
        from tools.slr_toolkit.llm_screening import run_llm_screening

        with pytest.raises(ValueError, match="delay"):
            run_llm_screening(delay=-0.1, estimate_only=True)
