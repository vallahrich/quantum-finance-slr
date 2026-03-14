"""Tests for search-run creation and search-log behavior."""

from __future__ import annotations

from openpyxl import load_workbook

from tools.slr_toolkit import config


def _search_log_ids(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    assert ws is not None
    ids = [
        str(value)
        for value, in ws.iter_rows(min_row=2, max_col=1, values_only=True)
        if value
    ]
    wb.close()
    return ids


class TestCreateSearchRun:
    def test_is_idempotent_for_search_log_rows(self, tmp_path, monkeypatch):
        from tools.slr_toolkit.search_run import create_search_run

        monkeypatch.setattr(config, "RAW_EXPORTS_DIR", tmp_path / "exports")
        monkeypatch.setattr(config, "SEARCH_LOG_XLSX", tmp_path / "search_log.xlsx")

        create_search_run("scopus", "2026-03-08")
        create_search_run("scopus", "2026-03-08")

        assert _search_log_ids(config.SEARCH_LOG_XLSX) == ["2026-03-08_scopus"]

    def test_can_skip_placeholder_search_log_entry(self, tmp_path, monkeypatch):
        from tools.slr_toolkit.search_run import create_search_run

        monkeypatch.setattr(config, "RAW_EXPORTS_DIR", tmp_path / "exports")
        monkeypatch.setattr(config, "SEARCH_LOG_XLSX", tmp_path / "search_log.xlsx")

        run_folder = create_search_run("openalex", "2026-03-08", log_search=False)

        assert run_folder.exists()
        assert not config.SEARCH_LOG_XLSX.exists()
